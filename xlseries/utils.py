import pandas as pd
import numpy as np
from openpyxl import load_workbook


def compare_cells(wb1, wb2):
    """Compare two excels based on row iteration."""

    # compare each cell of each worksheet
    for ws1, ws2 in zip(wb1.worksheets, wb2.worksheets):
        for row1, row2 in zip(ws1.iter_rows(), ws2.iter_rows()):
            for cell1, cell2 in zip(row1, row2):

                msg = "".join([_safe_str(cell1.value), " != ",
                               _safe_str(cell2.value), "row: ", str(cell1.row),
                               "column: ", str(cell1.column)])

                try:
                    value1 = float(cell1.value)
                    value2 = float(cell2.value)
                except:
                    value1 = cell1.value
                    value2 = cell2.value

                if type(value1) == float and type(value2) == float:
                    assert approx_equal(cell1.value, cell2.value, 0.00001), msg
                else:
                    assert cell1.value == cell2.value, msg
    return True


def _safe_str(value):

    if not value:
        RV = str(value)

    elif type(value) == str or type(value) == unicode:
        RV = value.encode("utf-8")

    else:
        RV = str(value)

    return RV


def approx_equal(a, b, tolerance):
    """Check if a and b can be considered approximately equal."""

    RV = False

    if (not a) and (not b):
        RV = True

    elif np.isnan(a) and np.isnan(b):
        # print a, type(a), "not approx_equal to", b, type(b)
        RV = True

    elif a and (a != np.nan) and b and (b != np.nan):
        RV = _approx_equal(a, b, tolerance)

    else:
        RV = a == b

    return RV


def _approx_equal(a, b, tolerance):
    if abs(a - b) < tolerance * a:
        return True
    else:
        return False


def infer_freq(av_seconds, tolerance=0.1):
    """Infer frequency of a time data series."""

    if approx_equal(1, av_seconds, tolerance):
        freq = 'S'
    elif approx_equal(60, av_seconds, tolerance):
        freq = 'T'
    elif approx_equal(3600, av_seconds, tolerance):
        freq = 'H'
    elif approx_equal(86400, av_seconds, tolerance):
        freq = 'D'
    elif approx_equal(604800, av_seconds, tolerance):
        freq = 'W'
    elif approx_equal(2419200, av_seconds, tolerance):
        freq = 'M'
    elif approx_equal(7776000, av_seconds, tolerance):
        freq = 'Q'
    elif approx_equal(15552000, av_seconds, tolerance):
        raise Exception("Can't handle semesters!")
    elif approx_equal(31536000, av_seconds, tolerance):
        freq = 'Y'
    else:
        raise Exception("Average seconds don't match any frequency.")

    return freq


def get_data_frames(xl_file):
    """Parse a well formatted excel file into pandas data frames."""

    dfs = []

    wb = load_workbook(filename=xl_file, use_iterators=True)
    ws_names = wb.get_sheet_names()

    for ws_name in ws_names:
        df = get_data_frame(xl_file, sheetname=ws_name)
        dfs.append(df)

    return dfs


def get_data_frame(xl_file, sheetname=0):
    """Parse a well formatted excel sheet into a pandas data frame."""

    df = pd.read_excel(xl_file, sheetname)

    # adopt a datetime index (first excel col)
    df = df.set_index(df.columns[0])

    time_delta = (df.index[-1] - df.index[0]) / df.index.size
    av_seconds = time_delta.total_seconds()
    period_range = pd.period_range(df.index[0],
                                   df.index[-1],
                                   freq=infer_freq(av_seconds))

    # rebuild data frame using a period range with frequency
    df = pd.DataFrame(data=df.values,
                      index=period_range,
                      columns=df.columns)

    return df
