#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
data_frame

Auxiliar methods to load and manipulate data frames.
"""
import pandas as pd
from openpyxl import load_workbook
import os
import arrow
import glob
import string

from .time_manipulation import infer_freq
from .comparing import approx_equal
from .xl_methods import normalize_value


class NoSerializedDataFrameFound(Exception):

    """Raises when no serialized data frame could be found."""
    pass


def get_data_frames(serial_df_path, use_period_range=True):
    """Parse a well formatted excel file into pandas data frames.

    Args:
        serial_dfs_path: Path to a serialized pandas data frame in .xlsx, .csv
            or .json format. If no extension is specified
            (eg. "the/path/to/test_case1" instead of
            "the/path/to/test_case1.xlsx") any of the former will be used.

    Returns:
        A list of pandas data frames.
    """

    dfs = []

    serial_df_path, extension = _parse_path_and_extension(serial_df_path)

    if extension == ".xlsx":
        wb = load_workbook(serial_df_path, read_only=True)
        ws_names = wb.sheetnames

        for ws_index in range(len(ws_names)):
            df = get_data_frame(serial_df_path, ws_index, use_period_range)
            dfs.append(df)

    elif extension == ".csv":
        if os.path.isfile(serial_df_path):
            dfs.append(get_data_frame(serial_df_path, use_period_range=True))

        elif os.path.isdir(serial_df_path.replace(".csv", "")):
            csv_dir = os.path.join(serial_df_path.replace(".csv", ""), "*.csv")

            for csv_file in glob.glob(csv_dir):
                dfs.append(get_data_frame(csv_file, use_period_range=True))

    elif extension == ".json":
        if os.path.isfile(serial_df_path):
            dfs.append(get_data_frame(serial_df_path, use_period_range=True))

        elif os.path.isdir(serial_df_path.replace(".json", "")):
            json_dir = os.path.join(serial_df_path.replace(".json", ""),
                                    "*.json")

            for json_file in glob.glob(json_dir):
                dfs.append(get_data_frame(json_file, use_period_range=True))

    return dfs


def get_data_frame(serial_df_path, index=0, use_period_range=True):
    """Parse a serialized pandas data frame.

    Serialized data frames are supported in xlsx, csv or json format. This can
    be specified in the path ("test_case1.xlsx") or left to this method to find
    any format available in the directory ("test_case1").

    Args:
        serial_df_path: Path to a serialized pandas data frame in .xlsx, .csv
            or .json format. If no extension is specified
            (eg. "the/path/to/test_case1" instead of
            "the/path/to/test_case1.xlsx") any of the former will be used.
        index: Index of the data frame to be loaded if the test case have more
            than one (this happens when a case has time series of different
            frequencies).
        use_period_range: Period range is a pandas index time that deals with
            ranges (eg "2013-04" for a monthly frequency) instead of precise
            datetime values. If false, datetime values (eg "2013-04-01") will
            be used.

    Returns:
        A pandas data frame.
    """

    serial_df_path, extension = _parse_path_and_extension(serial_df_path)

    # read df from the serialized file
    if extension == ".xlsx":
        df = pd.read_excel(serial_df_path, index)
        df = df.set_index(df.columns[0])

    elif extension == ".csv":
        if os.path.isdir(serial_df_path):
            letter = string.lowercase[index]
            base_name = os.path.basename(serial_df_path)
            file_name = base_name + letter + ".csv"
            serial_df_path = os.path.join(serial_df_path, file_name)

        df = pd.read_csv(serial_df_path)
        df = df.set_index(df.columns[0])

    elif extension == ".json":
        if os.path.isdir(serial_df_path):
            letter = string.lowercase[index]
            base_name = os.path.basename(serial_df_path)
            file_name = base_name + letter + ".json"
            serial_df_path = os.path.join(serial_df_path, file_name)

        elif (not os.path.isfile(serial_df_path) and
              os.path.isdir(serial_df_path[:-5])):

            letter = string.lowercase[index]
            base_name = os.path.basename(serial_df_path)
            file_name = base_name[:-5] + letter + ".json"
            serial_df_path = os.path.join(serial_df_path[:-5], file_name)

        with open(serial_df_path) as f:
            df = pd.read_json(f).sort_index()

    time_delta = ((arrow.get(df.index[-1]) - arrow.get(df.index[0])) /
                  df.index.size)
    av_seconds = time_delta.total_seconds()

    period_range = pd.date_range(df.index[0],
                                 df.index[-1],
                                 freq=infer_freq(av_seconds))

    # select time representation
    if use_period_range:
        time_index = period_range
    else:
        time_index = period_range.to_datetime()

    # rebuild data frame using a period range with frequency
    df = pd.DataFrame(data=df.values,
                      index=time_index,
                      columns=df.columns)

    return df


def _get_file_name_extension(file_name):
    index = file_name.rfind(".")
    if index == -1:
        return None
    else:
        return file_name[index:]


def _parse_path_and_extension(serial_df_path):
    supported_extensions = [".xlsx", ".csv", ".json"]
    extension = _get_file_name_extension(serial_df_path)
    serial_df_path_fixed = serial_df_path

    # check extension is valid or that there is a valid serial df of any ext
    if not extension or extension not in supported_extensions:
        for ext in supported_extensions:

            if os.path.isfile(serial_df_path + ext):
                serial_df_path_fixed = serial_df_path + ext
                extension = ext
                break

            elif os.path.isdir(serial_df_path):
                extension = ext
                break

        if not extension or extension not in supported_extensions:
            raise NoSerializedDataFrameFound(serial_df_path)

    return serial_df_path_fixed, extension


def dfs_to_json_and_csv(base_dir=os.getcwd()):
    """Convert data frames in excel files to json and csv format.

    Loads all excel files in a directory to data frames and dump them into
    additional json and csv files.

    Args:
        dir: Directory where all excel data frames are.
    """
    old_dir = os.getcwd()

    # safe check for Travis CI like build systems
    if not os.path.isdir(base_dir):
        base_dir = os.path.join(os.path.dirname(__file__), base_dir)

    os.chdir(base_dir)

    for test_case in glob.glob("*.xlsx"):
        dfs = get_data_frames(test_case, use_period_range=False)

        if len(dfs) == 1:
            dfs[0].to_json(test_case[:-5] + ".json")
            dfs[0].to_csv(test_case[:-5] + ".csv", encoding="utf-8")
        else:
            if not os.path.isdir(test_case[:-5]):
                os.mkdir(test_case[:-5])

            for index, df in enumerate(dfs):
                letter = string.lowercase[index]

                file_name = test_case[:-5] + letter + ".json"
                path = os.path.join(test_case[:-5], file_name)
                df.to_json(path)

                file_name = test_case[:-5] + letter + ".csv"
                path = os.path.join(test_case[:-5], file_name)
                df.to_csv(path, encoding="utf-8")

    os.chdir(old_dir)


def compare_data_frames(df1, df2):
    """Compare two data frames.

    An assertion is raised if data frames differ in index size, frequency,
    columns, index values or data values.

    Args:
        df1: First data frame to compare.
        df2: Second data frame to compare.

    Returns:
        True: When everything is the same in df1 and df2.

    Raises:
        AssertionError: When something is different between data frames.
    """
    assert isinstance(df1, pd.DataFrame), "df1 is not a DataFrame" + repr(df1)
    assert isinstance(df2, pd.DataFrame), "df2 is not a DataFrame" + repr(df2)

    msg = "Different index size"
    assert df1.index.size == df2.index.size, _diff_msg(msg, df1.index.size,
                                                       df2.index.size)

    msg = "Different index freq"
    assert df1.index.freqstr == df2.index.freqstr, _diff_msg(msg,
                                                             df1.index.freqstr,
                                                             df2.index.freqstr)

    msg = "Different columns"
    assert _check_columns(df1.columns, df2.columns), _diff_msg(msg,
                                                               df1.columns,
                                                               df2.columns)

    msg = "Different index"
    assert _check_index(df1.index, df2.index), _diff_msg(msg,
                                                         df1.index,
                                                         df2.index)

    msg = "Too different values"
    assert _check_values(df1.columns, df1, df2), msg

    return True


def _diff_msg(msg, elem1, elem2):
    """Creates a message for elements that differ in an assertion."""
    return msg + ": " + str(elem1) + " != " + str(elem2)


def _check_columns(cols1, cols2):
    """Check both column lists are equal."""

    for col1 in cols1:
        if col1 not in cols2:
            msg = "".join(["'", col1, "'", "\nnot in cols2\n",
                           "\n".join(list(cols2))])
            raise Exception(msg)

    for col2 in cols2:
        if col2 not in cols1:
            msg = "".join(["'", col2, "'", "\nnot in cols1\n",
                           "\n".join(list(cols1))])
            raise Exception(msg)

            return False

    return True


def _check_index(index1, index2):
    """Check two time indexes are equal."""

    for date1, date2 in zip(index1, index2):
        if not normalize_value(date1) == normalize_value(date2):
            return False

    return True


def _check_values(cols, df1, df2):
    """Check that all values of both data frames are approximately equal."""

    RV = True

    for col in cols:
        for value1, value2 in zip(df1[col], df2[col]):
            # print value1, value2, value2/value1-1
            if not approx_equal(value1, value2, 0.0001):
                print(value1, "and", value2, "not approx_equal")
                RV = False
                break

    return RV


def compare_period_ranges(pr1, pr2):
    """Compare two period ranges.

    Args:
        pr1: First period range to compare.
        pr2: Second period range to compare.

    Returns:
        Bool: True when period ranges are equal or False otherwise.
    """
    try:
        assert normalize_value(pr1.freq) == normalize_value(
            pr2.freq), "Different frequency"
        assert normalize_value(pr1[0]) == normalize_value(
            pr2[0]), "Different initial date"
        assert normalize_value(
            pr1[-1]) == normalize_value(pr2[-1]), "Different final date"

        return True

    except Exception as inst:
        print(inst)
        return False
