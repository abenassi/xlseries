#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
xl_methods

Useful methods for excel operations and related manipulations.
"""

from __future__ import print_function
from openpyxl import Workbook
from openpyxl.utils import column_index_from_string
import xlrd
import datetime
import pytz
import pandas
from .comparing import approx_equal


def common_row_or_column(coords_list):
    """Determine the common column or row index of a list of coords.

        Args:
            coords_list (list): List of consecutive coordinates.

        Returns:
            int: Number of the common row or column of the coordinates.
        """
    assert len(coords_list) >= 2, "There are less than 2 coords in the list."

    wb = Workbook()
    ws = wb.active

    row = ws[coords_list[0]].row
    col = ws[coords_list[0]].column

    if all([ws[coord].row == row for coord in coords_list]):
        return row

    elif all([ws[coord].column == col for coord in coords_list]):
        return column_index_from_string(col)

    else:
        raise Exception("There is no common row or column in " +
                        repr(coords_list))


def coord_in_scope(coord, coords):
    """Determine a coord is at the right or below a list of consecutive coords.

    Args:
        coord (str): A coordinate (eg "B1") than can be in the coords scope.
        coords (list): Coordinates (eg ["A1", "A2", "A3"] whose scope could
            include coord

    Returns:
        bool: True if shares a col or row with the list of coords that is not
            the common row or col of the list (can be both, if coord is one of
            the coordinates in the list of coords).
    """
    assert len(coords) >= 2, "There are less than 2 coords in the list."

    wb = Workbook()
    ws = wb.active

    row = ws[coords[0]].row
    col = ws[coords[0]].column

    if all([ws[scope_coord].row == row for scope_coord in coords]):
        return (ws[coord].row >= row and
                any([ws[scope_coord].column == ws[coord].column for
                     scope_coord in coords]))

    elif all([ws[scope_coord].column == col for scope_coord in coords]):
        return (ws[coord].column >= col and
                any([ws[scope_coord].row == ws[coord].row for
                     scope_coord in coords]))

    else:
        raise Exception("There is no common row or column in " +
                        repr(coords))


def consecutive_cells(cell_list):
    """True if cells are consecutive, False otherwise.

    Args:
        cell_list (list): List of strings that are coordinates.

    Example:
        >>> consecutive_cells(["A1", "A2", "A3"])
        True
        >>> consecutive_cells(["A1", "A2", "B2"])
        False
        >>> consecutive_cells(["A1", "B1", "C1"])
        True
        >>> consecutive_cells(["A1", "B1", "B2"])
        False
    """
    wb = Workbook()
    ws = wb.active

    row = None
    col = None
    alignment = None

    for cell in cell_list:

        if not row and not col:
            row = ws[cell].row
            col = column_index_from_string(ws[cell].column)

        elif not alignment:
            if ws[cell].row == row:
                alignment = "vertical"
                if col + 1 == column_index_from_string(ws[cell].column):
                    col += 1
                else:
                    return False

            elif column_index_from_string(ws[cell].column) == col:
                alignment = "horizontal"
                if row + 1 == ws[cell].row:
                    row += 1
                else:
                    return False

        else:
            if alignment == "vertical":
                if not ws[cell].row == row:
                    return False
                if not col + 1 == column_index_from_string(ws[cell].column):
                    return False
                col += 1

            else:
                if not column_index_from_string(ws[cell].column) == col:
                    return False
                if not row + 1 == ws[cell].row:
                    return False
                row += 1

    return True


def open_xls_as_xlsx(filename, data_only=True):
    """Open a xls file and return a openpyxl.Workbook.

    Args:
        filename: Path to an .xls file.

    Returns:
        Workbook: An openpyxl.Workbook.
    """
    assert filename[-4:] == ".xls", str(filename) + " is not an .xls file."

    wb_old = xlrd.open_workbook(filename)
    # TODO: data_only attribute must be changed because is deprecated
    # wb = Workbook(data_only=data_only)
    wb = Workbook()

    ws = wb.active
    wb.remove(ws)

    for ws_old in wb_old.sheets():
        index = 0
        nrows, ncols = 0, 0
        while nrows * ncols == 0:
            nrows = ws_old.nrows
            ncols = ws_old.ncols
            index += 1

        ws = wb.create_sheet(title=ws_old.name)

        for row in range(0, nrows):
            for col in range(0, ncols):
                ws.cell(row=row + 1, column=col + 1).value = ws_old.cell_value(
                    row, col)

    return wb


def make_wb_copy(wb):
    """Return a copy of an openpyxl workbook.

    Only taking into account sheet titles and cell values. Formatting is not
    being copied.

    Args:
        wb (Workbook): A workbook to make a copy from.

    Returns:
        Workbook: A copy made from wb.
    """
    wb_copy = Workbook()
    wb_copy.remove(wb_copy["Sheet"])

    for ws in wb:
        ws_copy = wb_copy.create_sheet(title=ws.title)
        for row in ws.rows:
            for cell in row:
                cell_copy = ws_copy[cell.column + str(cell.row)]
                cell_copy.value = cell.value

    return wb_copy


def make_ws_copy(ws):
    """Return a copy of an openpyxl worksheet.

    Only taking into account sheet titles and cell values. Formatting is not
    being copied.

    Args:
        ws (worksheet): A workbook to make a copy from.

    Returns:
        worksheet: A copy made from ws.
    """
    wb_copy = Workbook()
    wb_copy.remove(wb_copy["Sheet"])

    ws_copy = wb_copy.create_sheet(title=ws.title)
    for row in ws.rows:
        for cell in row:
            cell_copy = ws_copy[cell.column + str(cell.row)]
            cell_copy.value = cell.value

    return ws_copy


def xl_coordinates_range(start, end=None):
    """Creates a generator of excel coordinates.

    Args:
        start: Excel coordinate where range starts (eg. "A5").
        end: Excel coordinate where range ends (eg. "C7").

    Yields:
        A string with the coordinate that follows the previous one
            in a range. This are yielded row by row.

    >>> for coord in xl_coordinates_range("A1", "B2"):
    ...     print(coord)
    A1
    B1
    A2
    B2
    """

    ws = Workbook().active

    if end:
        for row in ws[start + ":" + end]:
            for cell in row:
                yield cell.coordinate
    else:
        yield start


def compare_cells(wb1, wb2):
    """Compare two excels based on row iteration."""

    # compare each cell of each worksheet
    for ws1, ws2 in zip(wb1.worksheets, wb2.worksheets):
        compare_cells_ws(ws1, ws2)
    return True


def compare_cells_ws(ws1, ws2):
    """Compare two worksheets based on row iteration."""

    # compare each cell of each worksheet
    for row1, row2 in zip(ws1.rows, ws2.rows):
        for cell1, cell2 in zip(row1, row2):

            msg = "".join([str(cell1.value), " != ",
                           str(cell2.value), "\nrow: ",
                           str(cell1.row),
                           " column: ", str(cell1.column)])

            try:
                value1 = float(cell1.value)
                value2 = float(cell2.value)

            except:
                value1 = normalize_value(cell1.value)
                value2 = normalize_value(cell2.value)

            if isinstance(value1, float) and isinstance(value2, float):
                assert approx_equal(cell1.value, cell2.value, 0.00001), msg
            else:
                assert value1 == value2, msg

    return True


def normalize_value(value):
    """Strip spaces if the value is a string, convert None to empty string or
    let it pass otherwise."""

    if isinstance(value, str):
        return value.strip()
    elif value is None:
        return ""
    elif isinstance(value, datetime.datetime):
        return value.replace(tzinfo=None)
    elif isinstance(value, pandas.Timestamp):
        return value.tz_localize(None)
    else:
        return value


def normalize_time_value(value):
    """Strip spaces if the value is a string, convert None to empty string or
    let it pass otherwise."""

    if isinstance(value, datetime.datetime):
        return value.replace(tzinfo=None)
    elif isinstance(value, pandas.Timestamp):
        return value.tz_localize(None)
    else:
        return value


def print_xl_range(ws, cells_range="A1:E10", width=15):
    """Print a representation of an excel cells range.

    Args:
        ws: Worksheet from where take the range.
        range: Cells range to print out.
    """

    for row in ws[cells_range]:

        for cell in row:
            value = str(cell.coordinate) + ": " + str(cell.value)

            # fix length of value representation
            if len(value) > width:
                value = value[:width - 3] + "..."
            else:
                value = value.ljust(width)

            # print new value
            print("| " + value, end=' ')

        # print last border of the row
        print("| ")

        # print the separator between rows
        for cell in row:
            print("| " + "-" * (width), end=' ')
        print("| ")


if __name__ == '__main__':
    import doctest
    doctest.testmod()
