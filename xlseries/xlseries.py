#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
xlseries

Main module to parse time data series inside excel files into Pandas
DataFrames. This is the only module that the user should use in the normal use
case.
"""

from __future__ import print_function

from openpyxl import load_workbook, Workbook
import imp
import os
import platform
from unidecode import unidecode

from .strategies import strategies
from .utils.xl_methods import make_wb_copy
from .strategies.discover.parameters import Parameters
from .utils.xl_methods import open_xls_as_xlsx
from .utils.path_finders import get_package_dir

import warnings
warnings.filterwarnings("ignore")


class XlSeries(object):

    """Time data series parser for excel files.

    Attributes:
        wb: Workbook object. The user can either pass the path where the excel
            file is located or the Workbook object with the xl already loaded.
    """

    def __init__(self, xl_path_or_wb):
        """Args:
            xl_path_or_wb (str or Workbook): Path to an excel file or a
                Workbook object.
        """
        self.xl_path_or_wb = xl_path_or_wb
        if isinstance(xl_path_or_wb, Workbook):
            self.wb = xl_path_or_wb
        else:
            self.wb = self._load_wb(xl_path_or_wb)
        self.params = {}

    @staticmethod
    def _load_wb(xl_path):
        """Load an xls or xlsx excel file.

        Args:
            xl_path (str): Path to an xls or xlsx file.

        Returns:
            Workbook: Loaded xl file in an openpyxl.Workbook object.
        """
        if xl_path[-5:] == ".xlsx":
            return load_workbook(xl_path, data_only=True)
        elif xl_path[-4:] == ".xls":
            return open_xls_as_xlsx(xl_path, data_only=True)
        else:
            raise ValueError(xl_path + " is not an .xls or .xlsx file.")

    # PUBLIC
    def get_data_frames(self, params_path_or_obj, ws_name=None,
                        safe_mode=False, preserve_wb_obj=True):
        """Scrape time series from an excel file into a pandas.DataFrame.

        Args:
            params_path_or_obj (str, dict or Parameters): Scraping parameters.
                str: Path to a JSON file with parameters.
                dict: Python dictionary with parameters like
                Parameters: A Parameters object already built.

            ws_name (str): Name of the worksheet that will be scraped.

            safe_mode (bool): When some parameters are not passed by the user,
                the safe mode will check all possible combinations, returning
                more than one result if many are found. If safe_mode is set to
                False, the first succesful result will be returned without
                checking the other possible combinations of parameters.

            preserve_wb_obj (bool): If True makes a safe copy of a workbook to
                preserve the original object without changes. Only use False if
                changes to the workbook object are not a problem.

        Returns:
            list: A list of pandas.DataFrame objects with time series scraped
                from the excel file. Every DataFrame in the list corresponds to
                a different frequency.

        Example:
            params = {"headers_coord": ["B1","C1"],
                      "data_starts": 2,
                      "frequency": "M",
                      "time_header_coord": "A1"}
            dfs = XlSeries(wb).get_data_frames(params)

        """
        # wb will be changed, so it has to be a copy to preserve the original
        if preserve_wb_obj:
            wb_copy = make_wb_copy(self.wb)
        else:
            wb_copy = self.wb
        ws_names = wb_copy.sheetnames

        if not ws_name:
            ws_name = ws_names[0]
            if len(ws_names) > 1:
                msg = "There are {} worksheets: {}\nThe first {} will be " + \
                    "analyzed"
                print(msg.format(len(ws_names),
                                 str([name.encode("utf-8")
                                      for name in ws_names]),
                                 ws_name.encode("utf-8")))
                print("Remember you can choose a different one passing a " + \
                    "ws_name keyword argument.")
        else:
            ws_name = self._sanitize_ws_name(ws_name, ws_names)

        for scraper in strategies.get_strategies():
            if scraper.accepts(wb_copy):
                scraper_obj = scraper(wb_copy, params_path_or_obj, ws_name)
                dfs, params = scraper_obj.get_data_frames(safe_mode)
                self.params[ws_name] = params

                if isinstance(dfs, list) and len(dfs) == 1:
                    return dfs[0]
                else:
                    return dfs

    @staticmethod
    def _sanitize_ws_name(ws_name_orig, ws_names):
        """Check the real ws name with certain tolerance to common mistakes."""

        if ws_name_orig in ws_names:
            return ws_name_orig

        elif unicode(ws_name_orig) in ws_names:
            return unicode(ws_name_orig)

        # check other ws names that may match
        else:
            for ws_name in ws_names:
                if unicode(ws_name_orig).strip() == unicode(ws_name).strip():
                    return ws_name

            for ws_name in ws_names:
                if unidecode(ws_name_orig).strip() == unidecode(ws_name).strip():
                    return ws_name

        return ws_name_orig

    @staticmethod
    def critical_params_template():
        """Return a template of critical params to fill and use.

        Returns:
            dict: A dictionary to fill with values.
        """
        return Parameters.get_critical_params_template()

    @staticmethod
    def complete_params_template():
        """Return a template of all the params to fill and use

        Returns:
            dict: A dictionary to fill with values.
        """
        return Parameters.get_complete_params_template()

    def open(self):
        """Open excel file with system's default program."""

        # save workbook if no path to excel file was given
        if isinstance(self.xl_path_or_wb, Workbook):
            filename = "temp_xl_file.xlsx"
            self.xl_path_or_wb.save(filename)
            path = filename
        else:
            path = self.xl_path_or_wb

        if platform.system().lower() == "windows":
            os.system(path)
        else:
            os.system("open " + path)
