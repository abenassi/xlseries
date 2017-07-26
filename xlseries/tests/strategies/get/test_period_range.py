#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
test_period_range

Tests for `period_range` module.
"""

from __future__ import unicode_literals
import unittest
import nose
from openpyxl import Workbook
import arrow
import pandas as pd

from xlseries.utils.data_frame import compare_period_ranges
from xlseries.strategies.get.period_range import GetPeriodRangesSingleFrequency
from xlseries.strategies.get.period_range import GetPeriodRangesMultifrequency


class GetPeriodRangesSingleFrequencyTestCase(unittest.TestCase):

    def test_get_period_ranges_vertical(self):
        test_class = GetPeriodRangesSingleFrequency
        wb = Workbook()
        ws = wb.active
        ws["A1"] = arrow.get(2000, 1, 1).datetime
        ws["A2"] = arrow.get(2000, 4, 1).datetime
        ws["A3"] = arrow.get(2000, 7, 1).datetime
        ws["A4"] = arrow.get(2000, 10, 1).datetime
        freq = "Q"
        ini_row = 1
        end_row = 4
        time_header_coord = "A1"
        time_alignement = 0
        alignment = "vertical"
        prs = test_class.get_period_ranges(ws, freq, ini_row,
                                           time_header_coord, end_row,
                                           time_alignement, alignment)
        pr_q = pd.date_range("20000101", "20001001", freq="QS")

        self.assertTrue(compare_period_ranges(pr_q, prs[0]))

    def test_get_period_ranges_horizontal(self):
        test_class = GetPeriodRangesSingleFrequency
        wb = Workbook()
        ws = wb.active
        ws["A1"] = arrow.get(2000, 1, 1).datetime
        ws["B1"] = arrow.get(2000, 4, 1).datetime
        ws["C1"] = arrow.get(2000, 7, 1).datetime
        ws["D1"] = arrow.get(2000, 10, 1).datetime
        freq = "Q"
        ini_col = 1
        end_col = 4
        time_header_coord = "A1"
        time_alignement = 0
        alignment = "horizontal"
        prs = test_class.get_period_ranges(ws, freq, ini_col,
                                           time_header_coord, end_col,
                                           time_alignement, alignment)
        pr_q = pd.date_range("20000101", "20001001", freq="QS")

        self.assertTrue(compare_period_ranges(pr_q, prs[0]))


class GetPeriodRangesMultifrequencyTestCase(unittest.TestCase):

    def test_get_period_ranges_vertical(self):
        test_class = GetPeriodRangesMultifrequency
        wb = Workbook()
        ws = wb.active
        ws["A1"] = arrow.get(2000, 1, 1).datetime
        ws["A2"] = arrow.get(2000, 1, 1).datetime
        ws["A3"] = arrow.get(2000, 4, 1).datetime
        ws["A4"] = arrow.get(2000, 7, 1).datetime
        ws["A5"] = arrow.get(2000, 10, 1).datetime
        ws["A6"] = arrow.get(2001, 1, 1).datetime
        ws["A7"] = arrow.get(2001, 1, 1).datetime
        ws["A8"] = arrow.get(2001, 4, 1).datetime
        ws["A9"] = arrow.get(2001, 7, 1).datetime
        freq = "AQQQQ"
        ini_row = 1
        end_row = 9
        time_header_coord = "A1"
        time_alignement = 0
        alignment = "vertical"
        prs = test_class.get_period_ranges(ws, freq, ini_row,
                                           time_header_coord, end_row,
                                           time_alignement, alignment)
        pr_y = pd.date_range("20000101", "20010101", freq="AS")
        pr_q = pd.date_range("20000101", "20010701", freq="QS")

        self.assertTrue(compare_period_ranges(pr_y, prs[0]))
        self.assertTrue(compare_period_ranges(pr_q, prs[1]))

        # complete the frequency
        ws["A10"] = arrow.get(2001, 10, 1).datetime
        end_row = 10

        prs = test_class.get_period_ranges(ws, freq, ini_row,
                                           time_header_coord, end_row,
                                           time_alignement, alignment)
        pr_y = pd.date_range("20000101", "20010101", freq="AS")
        pr_q = pd.date_range("20000101", "20011001", freq="QS")

        self.assertTrue(compare_period_ranges(pr_y, prs[0]))
        self.assertTrue(compare_period_ranges(pr_q, prs[1]))

    def test_get_period_ranges_horizontal(self):
        test_class = GetPeriodRangesMultifrequency
        wb = Workbook()
        ws = wb.active
        ws["A1"] = arrow.get(2000, 1, 1).datetime
        ws["B1"] = arrow.get(2000, 1, 1).datetime
        ws["C1"] = arrow.get(2000, 4, 1).datetime
        ws["D1"] = arrow.get(2000, 7, 1).datetime
        ws["E1"] = arrow.get(2000, 10, 1).datetime
        ws["F1"] = arrow.get(2001, 1, 1).datetime
        ws["G1"] = arrow.get(2001, 1, 1).datetime
        ws["H1"] = arrow.get(2001, 4, 1).datetime
        ws["I1"] = arrow.get(2001, 7, 1).datetime
        freq = "AQQQQ"
        ini_col = 1
        end_col = 9
        time_header_coord = "A1"
        time_alignement = 0
        alignment = "horizontal"
        prs = test_class.get_period_ranges(ws, freq, ini_col,
                                           time_header_coord, end_col,
                                           time_alignement, alignment)
        pr_y = pd.date_range("20000101", "20010101", freq="AS")
        pr_q = pd.date_range("20000101", "20010701", freq="QS")

        self.assertTrue(compare_period_ranges(pr_y, prs[0]))
        self.assertTrue(compare_period_ranges(pr_q, prs[1]))

        # complete the frequency
        ws["J1"] = arrow.get(2001, 10, 1).datetime
        end_col = 10

        prs = test_class.get_period_ranges(ws, freq, ini_col,
                                           time_header_coord, end_col,
                                           time_alignement, alignment)
        pr_y = pd.date_range("20000101", "20010101", freq="AS")
        pr_q = pd.date_range("20000101", "20011001", freq="QS")

        self.assertTrue(compare_period_ranges(pr_y, prs[0]))
        self.assertTrue(compare_period_ranges(pr_q, prs[1]))


if __name__ == '__main__':
    nose.run(defaultTest=__name__)
