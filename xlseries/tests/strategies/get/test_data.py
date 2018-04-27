#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
test_get_data_strategies

Tests for `get_data_strategies` module.
"""

import unittest
import nose
from openpyxl import Workbook
import arrow
import numpy as np

from xlseries.strategies.get.data import BaseAccepts
from xlseries.strategies.get.data import BaseGetDataStrategy
from xlseries.strategies.get.data import BaseSingleFrequency
from xlseries.strategies.get.data import BaseMultiFrequency
from xlseries.strategies.get.data import BaseContinuous
from xlseries.utils.comparing import compare_list_values

bases = (BaseAccepts, BaseSingleFrequency, BaseContinuous, BaseGetDataStrategy)
GetSingleFrequencyContinuous = type("CleanSingleColumn", bases, {})

bases = (BaseAccepts, BaseMultiFrequency, BaseContinuous, BaseGetDataStrategy)
GetMultiFrequencyContinuous = type("CleanSingleColumn", bases, {})


# @unittest.skip("skip")
class MissingsTestCase(unittest.TestCase):

    def _get_values(self, ws, ini_row, end_row, col):

        values = []
        i_row = ini_row
        while i_row <= end_row:
            values.append(ws.cell(row=i_row, column=col).value)
            i_row += 1

        return values

    def test_fill_implicit_missings_vertical(self):
        strategy = GetSingleFrequencyContinuous
        wb = Workbook()
        ws = wb.active

        ws["A1"] = arrow.get(2015, 6, 13).datetime
        ws["A2"] = arrow.get(2015, 6, 14).datetime
        ws["A3"] = arrow.get(2015, 6, 15).datetime
        ws["A4"] = arrow.get(2015, 6, 18).datetime
        ws["A5"] = arrow.get(2015, 6, 19).datetime
        ws["A6"] = arrow.get(2015, 6, 20).datetime
        ws["A7"] = arrow.get(2015, 6, 22).datetime
        ws["A8"] = arrow.get(2015, 6, 23).datetime

        values = list(range(8))
        frequency = "D"
        time_header_coord = "A1"
        ini_row = 1
        end_row = 8
        exp_values = [0, 1, 2, np.NaN, np.NaN, 3, 4, 5, np.NaN, 6, 7]

        new_values = strategy._fill_implicit_missings(ws, values, frequency,
                                                      time_header_coord,
                                                      ini_row,
                                                      end_row,
                                                      "vertical")

        self.assertEqual(len(new_values), len(exp_values))
        self.assertTrue(compare_list_values(new_values, exp_values))

    def test_fill_implicit_missings_horizontal(self):
        strategy = GetSingleFrequencyContinuous
        wb = Workbook()
        ws = wb.active

        ws["A1"] = arrow.get(2015, 6, 13).datetime
        ws["B1"] = arrow.get(2015, 6, 14).datetime
        ws["C1"] = arrow.get(2015, 6, 15).datetime
        ws["D1"] = arrow.get(2015, 6, 18).datetime
        ws["E1"] = arrow.get(2015, 6, 19).datetime
        ws["F1"] = arrow.get(2015, 6, 20).datetime
        ws["G1"] = arrow.get(2015, 6, 22).datetime
        ws["H1"] = arrow.get(2015, 6, 23).datetime

        values = list(range(8))
        frequency = "D"
        time_header_coord = "A1"
        ini_col = 1
        end_col = 8
        exp_values = [0, 1, 2, np.NaN, np.NaN, 3, 4, 5, np.NaN, 6, 7]

        new_values = strategy._fill_implicit_missings(ws, values, frequency,
                                                      time_header_coord,
                                                      ini_col,
                                                      end_col,
                                                      "horizontal")

        self.assertEqual(len(new_values), len(exp_values))
        self.assertTrue(compare_list_values(new_values, exp_values))


class BaseGetDataStrategyTestCase(unittest.TestCase):

    def setUp(self):
        self.base_class = BaseGetDataStrategy

        wb = Workbook()
        self.ws = wb.active

        self.ws["A1"] = "A"
        self.ws["A2"] = "B"
        self.ws["A3"] = "C"
        self.ws["A4"] = "GDP"

        self.ws["B1"] = "Agricultural"
        self.ws["B2"] = "Industrial"
        self.ws["B3"] = "Services"
        self.ws["B4"] = "Gross Domestic Product"

        self.ws["C6"] = 1

    def test_get_name_simple(self):
        name = self.base_class._get_name(self.ws, header_coord="B1")
        self.assertEqual(name, "Agricultural")

        name = self.base_class._get_name(self.ws, header_coord="C6")
        self.assertEqual(name, "1")

    def test_get_name_composed(self):
        name = self.base_class._get_name(self.ws, header_coord="B1",
                                         composed_headers_coord=["A1"])
        self.assertEqual(name, "A Agricultural")

    def test_get_name_context(self):
        name = self.base_class._get_name(self.ws, header_coord="B1",
                                         composed_headers_coord=["A1"],
                                         context=["GDP"])
        self.assertEqual(name, "GDP - A Agricultural")


if __name__ == '__main__':
    nose.run(defaultTest=__name__)
