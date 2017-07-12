#!/usr/bin/env python
# -*- coding: utf-8 -*-

import unittest
import nose
import os
import json
from functools import wraps
from openpyxl import Workbook

from xlseries.strategies.discover.parameters import Parameters
from xlseries.strategies.discover.parameters import InvalidParameter
from xlseries.strategies.discover.parameters import CriticalParameterMissing
from xlseries.utils.case_loaders import load_critical_parameters_case


"""
test_parameters

This module tests the parameters object.
"""


def get_orig_params_path(file_name):
    base_dir = os.path.dirname(__file__)
    return os.path.join(base_dir, "original", file_name)


def get_exp_params_path(file_name):
    base_dir = os.path.dirname(__file__)
    return os.path.join(base_dir, "expected", file_name)


# @unittest.skip("skip")
class ParametersTest(unittest.TestCase):

    def setUp(self):
        self.params = Parameters(get_orig_params_path("test_params.json"))
        self.params_exp = Parameters(get_exp_params_path("test_params.json"))

    def tearDown(self):
        del self.params

    # @unittest.skip("skip")
    def test_load_from_json(self):
        self.assertEqual(self.params.__dict__, self.params_exp.__dict__)

    def test_load_from_dict(self):
        with open(get_orig_params_path("test_params.json")) as f:
            params_dict = json.load(f)
        params = Parameters(params_dict)
        # pprint(params.__dict__)
        # pprint(self.params_exp.__dict__)

        for exp_params_name in self.params_exp.__dict__:
            self.assertEqual(params.__dict__[exp_params_name],
                             self.params_exp.__dict__[exp_params_name])

        for orig_params_name in params.__dict__:
            self.assertEqual(params.__dict__[orig_params_name],
                             self.params_exp.__dict__[orig_params_name])

    # @unittest.skip("skip")
    def test_get_num_series(self):
        self.assertEqual(self.params._get_num_series(self.params.__dict__), 3)
        self.assertEqual(self.params._get_num_series({"param": None}), None)

    def test_get_series_params(self):
        params = Parameters(get_orig_params_path(
            "test_params_time_multicolumn.json"))

        self.assertEqual(params["time_header_coord"], [["A1", "A2"],
                                                       ["A1", "A2"],
                                                       ["A1", "A2"]])

        self.assertEqual(params[0]["time_header_coord"], ["A1", "A2"])

    def test_valid_param_value(self):
        self.assertTrue(self.params._valid_param_value(True, [True, False]))
        self.assertTrue(self.params._valid_param_value(True, []))
        self.assertFalse(self.params._valid_param_value("A1", [True, False]))
        self.assertFalse(self.params._valid_param_value(None, [True, False]))

    def test_valid_freq(self):
        valid_freqs = ["Y", "Q", "M", "W", "D"]
        self.assertTrue(self.params._valid_freq("YQQQQ", valid_freqs))
        self.assertTrue(self.params._valid_freq("D", valid_freqs))
        self.assertFalse(self.params._valid_freq("YQQX", valid_freqs))

    def test_freq_translation(self):
        params = Parameters({
            "headers_coord": ["A1", "B1", "C1"],
            "data_starts": 2,
            "data_ends": 256,
            "frequency": "y",
            "time_header_coord": "A1",
        })
        self.assertEqual(params["frequency"], ["A", "A", "A"])

    def test_freq_translation(self):
        params = Parameters({
            "headers_coord": ["A1", "B1", "C1"],
            "data_starts": 2,
            "data_ends": 256,
            "frequency": "YQQQQ",
            "time_header_coord": "A1",
        })
        self.assertEqual(params["frequency"], ["AQQQQ", "AQQQQ", "AQQQQ"])

    def test_get_missings(self):
        params = Parameters({
            "alignment": None,
            "headers_coord": ["B1", "C1"],
            "data_starts": 2,
            "data_ends": 256,
            "frequency": "m",
            "time_header_coord": "A1",
            "time_multicolumn": None,
            "time_composed": None,
            "time_alignment": 0,
            "continuity": None,
            "blank_rows": None,
            "missings": None,
            "missing_value": None,
            "series_names": None,
            "composed_headers_coord": None,
            "context": None
        })
        exp_missings = ["time_composed", "continuity",
                        "blank_rows", "missings"]

        self.assertEqual(set(exp_missings), set(params.get_missings()))

    def test_validate_parameters_exception(self):
        params = {"continuity": "A1"}
        valid_values = {"continuity": [True, False]}
        with self.assertRaises(InvalidParameter):
            self.params._validate_parameters(params, valid_values)

    def test_remove_blank_headers(self):

        wb = Workbook()
        ws = wb.active

        params = Parameters({
            "headers_coord": ["A1", "B1", "C1"],
            "data_starts": 2,
            "data_ends": 256,
            "frequency": "m",
            "time_header_coord": "A1",
        })
        ws["A1"].value = "Importaciones"
        ws["B1"].value = "Exportaciones"
        params.remove_blank_headers(ws)

        self.assertEqual(params["headers_coord"], ["A1", "B1"])
        self.assertEqual(params["data_starts"], [2, 2])
        self.assertEqual(params["data_ends"], [256, 256])

        params = Parameters({
            "headers_coord": ["A1_A2", "B1", "C1_C2"],
            "data_starts": 2,
            "data_ends": 256,
            "frequency": "m",
            "time_header_coord": "A1",
        })
        ws["A1"].value = "Importaciones"
        ws["B1"].value = "Exportaciones"
        ws["C1"].value = "Saldo"
        params.remove_blank_headers(ws)

        self.assertEqual(params["headers_coord"], ["A2", "B1", "C2"])
        self.assertEqual(params["data_starts"], [2, 2, 2])
        self.assertEqual(params["data_ends"], [256, 256, 256])

        ws["E4"].value = "dont remove!"
        params = Parameters({
            "headers_coord": ["A1", "E1", "E2", "E3", "E4"],
            "data_starts": 2,
            "data_ends": 256,
            "frequency": "m",
            "time_header_coord": "A1",
        })
        ws["A1"].value = "Importaciones"
        ws["B1"].value = "Exportaciones"
        ws["C1"].value = "Saldo"
        params.remove_blank_headers(ws)

        self.assertEqual(params["headers_coord"], ["A1", "E4"])
        self.assertEqual(params["data_starts"], [2, 2])
        self.assertEqual(params["data_ends"], [256, 256])

    def test_remove_series(self):

        params = Parameters({
            "headers_coord": ["A1", "B1", "C1"],
            "data_starts": 2,
            "data_ends": 256,
            "frequency": "m",
            "time_header_coord": "A1",
        })
        params.remove_series(1)
        self.assertEqual(params.headers_coord, ["A1", "C1"])
        self.assertTrue(len(params.data_starts), 2)
        self.assertTrue(len(params.time_header_coord), 2)


class ParametersClassMethodsTest(unittest.TestCase):

    """Test class mehtods that don't need to load parameters to be tested."""

    def test_ensure_critical_parameters_exception(self):
        params = {"data_starts": None}
        critical = ["data_starts"]
        valid_values = {"data_starts": [int]}
        with self.assertRaises(CriticalParameterMissing):
            Parameters._check_has_critical(params, critical,
                                           valid_values)

    def test_check_consistency(self):
        params_dict = {"data_starts": 1,
                       "headers_coord": ["A2", "B2", "C2", "D2"]}
        with self.assertRaises(AssertionError):
            Parameters._check_consistency(params_dict)

        params_dict = {"data_starts": 1,
                       "headers_coord": ["B1", "B2", "B3", "B4"]}
        with self.assertRaises(AssertionError):
            Parameters._check_consistency(params_dict)

    def test_guess_alignment(self):
        headers = ["A1", "B1", "C1"]
        self.assertEqual(Parameters._guess_alignment(headers), "vertical")

        headers = ["A1", "B1", "D1", "E1"]
        self.assertEqual(Parameters._guess_alignment(headers), "vertical")

        headers = ["A1", "A2"]
        self.assertEqual(Parameters._guess_alignment(headers), "horizontal")

        headers = ["A1", "A3", "A5", "A7"]
        self.assertEqual(Parameters._guess_alignment(headers), "horizontal")

        headers = ["A1", "A3", "A5"]
        self.assertEqual(Parameters._guess_alignment(headers), None)

        headers = ["A1", "A3", "A5", "B7"]
        self.assertEqual(Parameters._guess_alignment(headers), None)

    def test_apply_to_all_missing_value(self):
        missing_value = "-"
        num_series = 3
        res = Parameters._apply_to_all_missing_value(missing_value, num_series)
        exp = [["-"], ["-"], ["-"]]
        self.assertEqual(res, exp)

        missing_value = ["-"]
        num_series = 3
        res = Parameters._apply_to_all_missing_value(missing_value, num_series)
        exp = [["-"], ["-"], ["-"]]
        self.assertEqual(res, exp)

        missing_value = ["-", "."]
        num_series = 3
        res = Parameters._apply_to_all_missing_value(missing_value, num_series)
        exp = [["-", "."], ["-", "."], ["-", "."]]
        self.assertEqual(res, exp)

        missing_value = []
        num_series = 3
        res = Parameters._apply_to_all_missing_value(missing_value, num_series)
        exp = [[], [], []]
        self.assertEqual(res, exp)

    def test_unpack_header_ranges(self):

        exp = ["A5", "A6", "A7", "A8"]
        self.assertEqual(Parameters._unpack_header_ranges("a5-A8"), exp)

        exp = ["A5", "B5", "C5"]
        self.assertEqual(Parameters._unpack_header_ranges("A5-c5"), exp)

        exp = ["A5"]
        self.assertEqual(Parameters._unpack_header_ranges("a5"), exp)

        exp = None
        self.assertEqual(Parameters._unpack_header_ranges("None"), exp)

        exp = [["A1", "A2"], ["A1", "A2"]]
        orig = [["A1", "A2"], ["A1", "A2"]]
        self.assertEqual(Parameters._unpack_header_ranges(orig), exp)

        exp = [["A1", "A2", "A3"], ["A1", "A2", "A3"]]
        orig = [["A1-A3"], ["A1-A3"]]
        self.assertEqual(Parameters._unpack_header_ranges(orig), exp)

    def test_unpack_composed_header_ranges(self):

        exp = ["A5_B5", "A6_B6", "A7_B7", "A8_B8"]
        self.assertEqual(
            Parameters._unpack_header_ranges("(a5_B5)-(A8_B8)"), exp)

        exp = [["A1_B1", "A2_B2", "A3_B3"], ["A1_B1", "A2_B2", "A3_B3"]]
        orig = [["(A1_B1)-(A3_b3)"], ["(A1_B1)-(A3_b3)"]]
        self.assertEqual(Parameters._unpack_header_ranges(orig), exp)

    def test_separate_composed_headers(self):

        headers_coord = ["A1_B1", "A2_B2", "A3_B3"]
        exp = ([["A1"], ["A2"], ["A3"]], ["B1", "B2", "B3"])
        self.assertEqual(Parameters._separate_composed_headers(headers_coord),
                         exp)

        headers_coord = ["A1_B1_C1", "A2_B2_C2", "A3_B3_C3"]
        exp = ([["A1", "B1"], ["A2", "B2"], ["A3", "B3"]], ["C1", "C2", "C3"])
        self.assertEqual(Parameters._separate_composed_headers(headers_coord),
                         exp)

    def test_process_context(self):

        context = {"GDP": "A5-A8"}
        headers_coord = ["A5", "A6", "A7", "A8"]
        exp_context = [["GDP"], ["GDP"], ["GDP"], ["GDP"]]
        self.assertEqual(Parameters._process_context(context, headers_coord),
                         exp_context)

        context = {"GDP": "A5-A8"}
        headers_coord = ["B5", "B6", "B7", "B8"]
        exp_context = [["GDP"], ["GDP"], ["GDP"], ["GDP"]]
        self.assertEqual(Parameters._process_context(context, headers_coord),
                         exp_context)

        context = {"GDP": ["A5-A6", "A8-A9"]}
        headers_coord = ["B5", "B6", "B7", "B8", "B9"]
        exp_context = [["GDP"], ["GDP"], [], ["GDP"], ["GDP"]]
        self.assertEqual(Parameters._process_context(context, headers_coord),
                         exp_context)

        context = {"GDP": ["A5-A6", "A8-A9"],
                   "Agricultural": "A5-A6",
                   "Industrial": "A8-A9"}
        headers_coord = ["B5", "B6", "B7", "B8", "B9"]
        exp_context = [["GDP", "Agricultural"], ["GDP", "Agricultural"], [],
                       ["GDP", "Industrial"], ["GDP", "Industrial"]]
        self.assertEqual(Parameters._process_context(context, headers_coord),
                         exp_context)


def load_case_number():
    """Decorate a test loading the case number taken from test name."""

    def fn_decorator(fn):
        case_num = int(fn.__name__.split("_")[1][-1])

        @wraps(fn)
        def fn_decorated(*args, **kwargs):
            kwargs["case_num"] = case_num
            fn(*args, **kwargs)

        return fn_decorated
    return fn_decorator


# @unittest.skip("skip")
class ParametersCriticalDictTestCase(unittest.TestCase):

    """Test Parameters loading dict with only critical parameters."""

    CRITICAL_PARAMS = {

        1: {'data_starts': 2,
            'frequency': u'm',
            'headers_coord': [u'B1', u'C1'],
            'time_header_coord': u'A1'},

        2: {'blank_rows': [False, True],
            'continuity': [True, False],
            'data_starts': [5, 22],
            'frequency': [u'D', u'M'],
            'headers_coord': [u'D4', u'F4'],
            'missing_value': [u'Implicit', None],
            'missings': [True, False],
            'time_alignment': [0, -1],
            'time_header_coord': [u'C4', u'F4']},

        3: {'data_starts': 7,
            'frequency': u'Q',
            'headers_coord': [u'B4', u'C4', u'D4'],
            'time_header_coord': u'A4'},

        4: {'data_starts': [5, 5, 5, 5, 52, 52, 52, 52],
            'frequency': u'q',
            'headers_coord': [u'B4', u'C4', u'D4', u'E4', u'B51', u'C51',
                              u'D51', u'E51'],
            'time_header_coord': [u'A4', u'A4', u'A4', u'A4', u'A51', u'A51',
                                  u'A51', u'A51'],
            'missing_value': [u'\u2026']},

        5: {'data_starts': 28,
            'frequency': u'M',
            'headers_coord': [u'G22', u'H22'],
            'time_header_coord': u'A18'},

        6: {'data_starts': 3,
            'frequency': u'yQQQQ',
            'headers_coord': ['B8', 'B9', 'B10', 'B11', 'B12', 'B13', 'B14',
                              'B15', 'B16', 'B17', 'B18', 'B19', 'B20', 'B21',
                              'B22', 'B23', 'B24', 'B25', 'B26', 'B27', 'B28'],
            'time_header_coord': [u'C4', u'C6']},

        7: {'data_starts': 2,
            'frequency': u'Y',
            'headers_coord': [u'A8', 'A10', 'A11', 'A12', 'A14', 'A15', 'A16',
                              'A18', 'A19', 'A20', 'A21', 'A22', 'A24',
                              'A25', 'A26', u'A28', u'A30', u'A32', u'A34',
                              'A36', 'A37', 'A38', 'A39', 'A41', 'A42',
                              'A43', 'A44', u'A46', u'A48', 'A50', 'A51',
                              'A52', u'A55'],
            'time_header_coord': u'A6'}
    }

    def check_critical_dict_params(self, case_num):
        """Check critical dict parameters loading.

        Args:
            case_num (int): The test case number to run.
        """

        params = Parameters(self.CRITICAL_PARAMS[case_num].copy())
        exp_params = load_critical_parameters_case(case_num)

        # override the guessing of Parameters
        params.remove("alignment")

        self.assertEqual(params, exp_params)

    # @unittest.skip("skip")
    @load_case_number()
    def test_case1(self, case_num):
        self.check_critical_dict_params(case_num)

    # @unittest.skip("skip")
    @load_case_number()
    def test_case2(self, case_num):
        self.check_critical_dict_params(case_num)

    # @unittest.skip("skip")
    @load_case_number()
    def test_case3(self, case_num):
        self.check_critical_dict_params(case_num)

    # @unittest.skip("skip")
    @load_case_number()
    def test_case4(self, case_num):
        self.check_critical_dict_params(case_num)

    # @unittest.skip("skip")
    @load_case_number()
    def test_case5(self, case_num):
        self.check_critical_dict_params(case_num)

    # @unittest.skip("skip")
    @load_case_number()
    def test_case6(self, case_num):
        self.check_critical_dict_params(case_num)

    # @unittest.skip("skip")
    @load_case_number()
    def test_case7(self, case_num):
        self.check_critical_dict_params(case_num)

    # @unittest.skip("skip")
    def test_case_external1(self):
        p = {'data_starts': 2,
             'frequency': 'Q',
             'headers_coord': 'A53',
             'time_header_coord': 'A52'}
        params = Parameters(p)
        self.assertTrue(params["alignment"] is None)

        p2 = {'alignment': 'horizontal',
              'data_starts': 2,
              'frequency': 'Q',
              'headers_coord': 'A53',
              'time_header_coord': 'A52'}
        params = Parameters(p2)
        self.assertEqual(params["alignment"][0], "horizontal")


# @unittest.skip("skip")
class ParametersIntegrationTestCase(unittest.TestCase):

    """Test Parameters loading sets of user inputed parameters.

    Check that the sanitization and completion of the user inputed sets
    of parameters is the one that should be expected."""

    def test_composed_headers(self):
        p = {'data_starts': 4,
             'frequency': 'Q',
             'headers_coord': '(A2_B2)-(A5_B5)',
             'time_header_coord': 'A1'}
        params = Parameters(p)

        exp_headers_coord = ["B2", "B3", "B4", "B5"]
        self.assertEqual(params["headers_coord"], exp_headers_coord)

        exp_composed_headers_coord = [["A2"], ["A3"], ["A4"], ["A5"]]
        self.assertEqual(params["composed_headers_coord"],
                         exp_composed_headers_coord)

    def test_context(self):
        p = {'data_starts': 2,
             'frequency': 'Q',
             'headers_coord': 'A2-A5',
             'time_header_coord': 'A1',
             'context': {"GDP": "A2-A5"}}
        params = Parameters(p)

        exp_context = [["GDP"], ["GDP"], ["GDP"], ["GDP"]]
        self.assertEqual(params["context"], exp_context)

if __name__ == '__main__':
    nose.run(defaultTest=__name__)
