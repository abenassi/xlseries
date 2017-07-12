#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
test_clean_ti_strategies

Tests for `clean_ti_strategies` module.
"""

import unittest
import nose
import arrow
import datetime
import os
from openpyxl import load_workbook, Workbook

from xlseries.strategies.clean.time_index import BaseCleanTiStrategy
from xlseries.strategies.clean.time_index import BaseAccepts
from xlseries.strategies.clean.time_index import BaseNoOffsetTi
from xlseries.strategies.clean.time_index import BaseSingleTable
from xlseries.strategies.clean.time_index import BaseSingleColumn
from xlseries.strategies.clean.time_index import BaseMultipleColumns
from xlseries.strategies.clean.time_index import BaseSingleFrequency
from xlseries.strategies.clean.time_index import BaseMultiFrequency
from xlseries.strategies.clean.time_index import TimeValueGoingBackwards
from xlseries.strategies.clean.time_index import TimeValueGoingForth
from xlseries.utils.xl_methods import compare_cells
from xlseries.utils.case_loaders import load_parameters_case
from xlseries.utils.path_finders import abs_path


bases = (BaseAccepts, BaseSingleTable, BaseSingleColumn, BaseSingleFrequency,
         BaseNoOffsetTi, BaseCleanTiStrategy)
CleanSingleColumn = type("CleanSingleColumn", bases, {})

bases = (BaseAccepts, BaseSingleTable, BaseMultipleColumns,
         BaseSingleFrequency, BaseNoOffsetTi, BaseCleanTiStrategy)
CleanMultipleColumns = type("CleanMultipleColumns", bases, {})

bases = (BaseAccepts, BaseSingleTable, BaseMultipleColumns, BaseMultiFrequency,
         BaseNoOffsetTi, BaseCleanTiStrategy)
CleanMultiColumnsMultiFreq = type("CleanMultiColumnsMultiFreq", bases, {})


# @unittest.skip("skip")
class BaseCleanTiStrategyTestCase(unittest.TestCase):

    def test_correct_progression_backwards_exception(self):
        last = arrow.get(2015, 5, 9)
        curr = arrow.get(2015, 2, 15)
        freq = "D"
        missings = False

        with self.assertRaises(TimeValueGoingBackwards):
            BaseCleanTiStrategy._correct_progression(last, curr, freq,
                                                     missings)

    def test_correct_progression_forth_exception(self):
        last = arrow.get(2015, 5, 9)
        curr = arrow.get(2016, 6, 30)
        freq = "M"
        missings = False

        with self.assertRaises(TimeValueGoingForth):
            BaseCleanTiStrategy._correct_progression(last, curr, freq,
                                                     missings)


# @unittest.skip("skip")
class CleanSingleColumnTestCase(unittest.TestCase):

    # @unittest.skip("skip")

    def test_time_index_iterator(self):

        wb = Workbook()
        ws = wb.active

        ws["A1"].value = "a"
        ws["A2"].value = "b"
        ws["A3"].value = "c"

        alignment = "vertical"
        time_header_coord = "A1"
        ini = 1
        end = 3

        ti_iter = CleanSingleColumn._time_index_iterator(ws, alignment,
                                                         time_header_coord,
                                                         ini, end)
        res = [i[0] for i in ti_iter]
        self.assertEqual(res, ["a", "b", "c"])

        ti_iter = CleanSingleColumn._time_index_iterator(ws, alignment,
                                                         time_header_coord,
                                                         ini)
        res = [i[0] for i in ti_iter]
        self.assertEqual(res, ["a", "b", "c", None])

        ws["F1"].value = "d"
        ws["G1"].value = "e"
        ws["H1"].value = "f"

        alignment = "horizontal"
        time_header_coord = "F1"
        ini = 6
        end = 8
        ti_iter = CleanSingleColumn._time_index_iterator(ws, alignment,
                                                         time_header_coord,
                                                         ini, end)
        res = [i[0] for i in ti_iter]
        self.assertEqual(res, ["d", "e", "f"])

        ti_iter = CleanSingleColumn._time_index_iterator(ws, alignment,
                                                         time_header_coord,
                                                         ini)
        res = [i[0] for i in ti_iter]
        self.assertEqual(res, ["d", "e", "f", None])

    def test_correct_progression(self):

        # progression wrong because going to the past
        last_time_value = arrow.get(2011, 7, 5)
        curr_time_value = arrow.get(2011, 5, 6)
        freq = "D"
        missings = True
        missing_value = "Implicit"

        new_time_value = CleanSingleColumn()._correct_progression(
            last_time_value,
            curr_time_value,
            freq, missings,
            missing_value)
        exp_time_value = arrow.get(2011, 7, 6)

        self.assertEqual(new_time_value, exp_time_value)

        # progression wrong because going to the future
        curr_time_value = arrow.get(2011, 8, 6)
        new_time_value = CleanSingleColumn()._correct_progression(
            last_time_value,
            curr_time_value,
            freq, missings,
            missing_value)

        self.assertEqual(new_time_value, exp_time_value)

    # @unittest.skip("skip")
    def test_parse_time(self):

        value = "17-12.09"
        last_time = arrow.get(2009, 12, 16)

        params = load_parameters_case(2)
        # print repr(params[0])

        new_time_value = CleanSingleColumn()._parse_time(params[0], value,
                                                         last_time)

        exp_time_value = arrow.get(2009, 12, 17)

        self.assertEqual(new_time_value, exp_time_value)

    # @unittest.skip("skip")
    def test_clean_time_index_case3(self):

        wb = load_workbook(
            os.path.join(abs_path("original"), "test_case3.xlsx"))
        ws = wb.active

        params = {"alignment": "vertical",
                  "time_alignment": 0,
                  "continuity": True,
                  "blank_rows": False,
                  "time_header_coord": "A4",
                  "data_starts": 7,
                  "data_ends": 119,
                  "frequency": "Q",
                  "missings": False,
                  "missing_value": None,
                  "time_multicolumn": False,
                  "time_composed": True}

        CleanSingleColumn().clean_time_index(ws, params)

        wb_exp = load_workbook(
            os.path.join(abs_path("expected"), "test_case3.xlsx"))

        # wb.save("test_case2_after_cleaning_index.xlsx")
        self.assertTrue(compare_cells(wb, wb_exp))

    # @unittest.skip("skip")
    def test_clean_time_index_case3_without_end(self):

        wb = load_workbook(
            os.path.join(abs_path("original"), "test_case3.xlsx"))
        ws = wb.active

        params = {"alignment": "vertical",
                  "time_alignment": 0,
                  "continuity": True,
                  "blank_rows": False,
                  "time_header_coord": "A4",
                  "data_starts": 7,
                  "data_ends": None,
                  "frequency": "Q",
                  "missings": False,
                  "missing_value": None,
                  "time_multicolumn": False,
                  "time_composed": True}

        end = CleanSingleColumn().clean_time_index(ws, params)

        wb_exp = load_workbook(
            os.path.join(abs_path("expected"), "test_case3.xlsx"))

        # wb.save("test_case2_after_cleaning_index.xlsx")
        self.assertTrue(compare_cells(wb, wb_exp))
        self.assertEqual(end, 119)

    # @unittest.skip("skip")
    def test_clean_time_index_case1(self):

        wb = load_workbook(
            os.path.join(abs_path("original"), "test_case1.xlsx"))
        ws = wb.active

        params = {"alignment": "vertical",
                  "time_alignment": 0,
                  "continuity": True,
                  "blank_rows": False,
                  "time_header_coord": "A1",
                  "data_starts": 2,
                  "data_ends": 256,
                  "frequency": "M",
                  "missings": True,
                  "missing_value": None,
                  "time_multicolumn": False,
                  "time_composed": False}

        CleanSingleColumn().clean_time_index(ws, params)

        wb_exp = load_workbook(
            os.path.join(abs_path("expected"), "test_case1.xlsx"))

        # wb.save("test_case2_after_cleaning_index.xlsx")
        self.assertTrue(compare_cells(wb, wb_exp))

    # @unittest.skip("skip")
    def test_clean_time_index_case1_without_end(self):

        wb = load_workbook(
            os.path.join(abs_path("original"), "test_case1.xlsx"))
        ws = wb.active

        params = {"alignment": "vertical",
                  "time_alignment": 0,
                  "continuity": True,
                  "blank_rows": False,
                  "time_header_coord": "A1",
                  "data_starts": 2,
                  "data_ends": None,
                  "frequency": "M",
                  "missings": True,
                  "missing_value": None,
                  "time_multicolumn": False,
                  "time_composed": False}

        end = CleanSingleColumn().clean_time_index(ws, params)

        wb_exp = load_workbook(
            os.path.join(abs_path("expected"), "test_case1.xlsx"))

        # wb.save("test_case2_after_cleaning_index.xlsx")
        self.assertTrue(compare_cells(wb, wb_exp))
        self.assertEqual(end, 256)

    # @unittest.skip("skip")
    def test_clean_time_index_case2(self):

        wb = load_workbook(
            os.path.join(abs_path("original"), "test_case2.xlsx"))
        ws = wb.active

        params = {"alignment": "vertical",
                  "time_alignment": 0,
                  "time_format": datetime.datetime,
                  "continuity": True,
                  "blank_rows": True,
                  "time_header_coord": "C4",
                  "data_starts": 5,
                  "data_ends": 2993,
                  "frequency": "D",
                  "missings": True,
                  "missing_value": "Implicit",
                  "time_multicolumn": False,
                  "time_composed": False}

        CleanSingleColumn().clean_time_index(ws, params)

        wb_exp = load_workbook(
            os.path.join(abs_path("expected"), "test_case2.xlsx"))

        # wb.save("test_case2_after_cleaning_index.xlsx")
        self.assertTrue(compare_cells(wb, wb_exp))

    def test_clean_time_index_case2_without_end(self):

        wb = load_workbook(
            os.path.join(abs_path("original"), "test_case2.xlsx"))
        ws = wb.active

        params = {"alignment": "vertical",
                  "time_alignment": 0,
                  "time_format": datetime.datetime,
                  "continuity": True,
                  "blank_rows": True,
                  "time_header_coord": "C4",
                  "data_starts": 5,
                  "data_ends": None,
                  "frequency": "D",
                  "missings": True,
                  "missing_value": "Implicit",
                  "time_multicolumn": False,
                  "time_composed": False}

        end = CleanSingleColumn().clean_time_index(ws, params)

        wb_exp = load_workbook(
            os.path.join(abs_path("expected"), "test_case2.xlsx"))

        # wb.save("test_case2_after_cleaning_index.xlsx")
        self.assertTrue(compare_cells(wb, wb_exp))
        self.assertEqual(end, 2993)

    # @unittest.skip("skip")
    def test_clean_time_index_case5(self):

        wb = load_workbook(
            os.path.join(abs_path("original"), "test_case5.xlsx"))
        ws = wb.active

        params = {"alignment": "vertical",
                  "time_alignment": 0,
                  "time_format": str,
                  "continuity": False,
                  "blank_rows": True,
                  "time_header_coord": "A18",
                  "data_starts": 28,
                  "data_ends": 993,
                  "frequency": "M",
                  "missings": True,
                  "missing_value": None,
                  "time_multicolumn": False,
                  "time_composed": True}

        CleanSingleColumn().clean_time_index(ws, params)

        wb_exp = load_workbook(
            os.path.join(abs_path("expected"), "test_case5.xlsx"))

        # wb.save("test_case5_after_cleaning_index.xlsx")
        self.assertTrue(compare_cells(wb, wb_exp))

    # @unittest.skip("skip")
    def test_clean_time_index_case5_without_end(self):

        wb = load_workbook(
            os.path.join(abs_path("original"), "test_case5.xlsx"))
        ws = wb.active

        params = {"alignment": "vertical",
                  "time_alignment": 0,
                  "time_format": str,
                  "continuity": False,
                  "blank_rows": True,
                  "time_header_coord": "A18",
                  "data_starts": 28,
                  "data_ends": None,
                  "frequency": "M",
                  "missings": True,
                  "missing_value": None,
                  "time_multicolumn": False,
                  "time_composed": True}

        end = CleanSingleColumn().clean_time_index(ws, params)

        wb_exp = load_workbook(
            os.path.join(abs_path("expected"), "test_case5_without_end.xlsx"))

        # wb.save("test_case5_after_cleaning_index.xlsx")
        self.assertTrue(compare_cells(wb, wb_exp))
        self.assertEqual(end, 1001)

    # @unittest.skip("skip")
    def test_clean_time_index_case7(self):

        wb = load_workbook(
            os.path.join(abs_path("original"), "test_case7.xlsx"))
        ws = wb.active

        params = {"alignment": "horizontal",
                  "blank_rows": True,
                  "composed_headers": False,
                  "data_starts": 2,
                  "data_ends": 44,
                  "frequency": "A",
                  "continuity": False,
                  "missings": False,
                  "missing_value": "None",
                  "multifrequency": False,
                  "series_names": "None",
                  "time_composed": True,
                  "time_alignment": 0,
                  "time_multicolumn": False,
                  "time_format": int,
                  "time_header": False,
                  "time_header_coord": "A6"}

        CleanSingleColumn().clean_time_index(ws, params)

        wb_exp = load_workbook(
            os.path.join(abs_path("expected"), "test_case7.xlsx"),
            data_only=True)

        # wb.save("test_case5_after_cleaning_index.xlsx")
        self.assertTrue(compare_cells(wb, wb_exp))

    def test_clean_time_index_case7_without_end(self):

        wb = load_workbook(
            os.path.join(abs_path("original"), "test_case7.xlsx"))
        ws = wb.active

        params = {"alignment": "horizontal",
                  "blank_rows": True,
                  "composed_headers": False,
                  "data_starts": 2,
                  "data_ends": None,
                  "frequency": "A",
                  "continuity": False,
                  "missings": False,
                  "missing_value": "None",
                  "multifrequency": False,
                  "series_names": "None",
                  "time_composed": True,
                  "time_alignment": 0,
                  "time_multicolumn": False,
                  "time_format": int,
                  "time_header": False,
                  "time_header_coord": "A6"}

        end = CleanSingleColumn().clean_time_index(ws, params)

        wb_exp = load_workbook(
            os.path.join(abs_path("expected"), "test_case7.xlsx"),
            data_only=True)

        # wb.save("test_case5_after_cleaning_index.xlsx")
        self.assertTrue(compare_cells(wb, wb_exp))
        self.assertEqual(end, 44)

    # @unittest.skip("skip")
    def test_forth_time_value_typo(self):

        exp_time = arrow.get(2015, 5, 2)
        max_forth_time = arrow.get(2015, 5, 22)
        curr_time = arrow.get(2015, 7, 2)
        fixed_time = BaseCleanTiStrategy._forth_time_value_typo(curr_time,
                                                                max_forth_time)
        self.assertEqual(exp_time, fixed_time)


# @unittest.skip("skip")
class CleanMultipleColumnsTestCase(unittest.TestCase):

    # @unittest.skip("skip")

    def test_clean_time_index_case5b(self):

        wb = load_workbook(
            os.path.join(abs_path("original"), "test_case5b.xlsx"))
        ws = wb.active

        params = {"alignment": "vertical",
                  "time_alignment": 0,
                  "time_format": str,
                  "time_header_coord": ["A18", "B18"],
                  "data_starts": 28,
                  "data_ends": 993,
                  "frequency": "M",
                  "missings": True,
                  "missing_value": None,
                  "time_multicolumn": True,
                  "continuity": False,
                  "time_composed": True}

        CleanMultipleColumns()._clean_time_index(ws, params)

        wb_exp = load_workbook(
            os.path.join(abs_path("expected"), "test_case5.xlsx"))

        # wb.save("test_case5b_after_cleaning_index.xlsx")
        self.assertTrue(compare_cells(wb, wb_exp))

    def test_time_index_iterator(self):

        wb = Workbook()
        ws = wb.active

        ws["A1"].value = "a"
        ws["A2"].value = "b"
        ws["A3"].value = "c"
        ws["B1"].value = "1"
        ws["B2"].value = "2"
        ws["B3"].value = "3"

        alignment = "vertical"
        time_header_coord = ["A1", "B1"]
        ini = 1
        end = 3
        ti_iter = CleanMultipleColumns._time_index_iterator(ws, alignment,
                                                            time_header_coord,
                                                            ini, end)
        res = [i[0] for i in ti_iter]
        self.assertEqual(res, ["a 1", "b 2", "c 3"])

        ws["F1"].value = "d"
        ws["G1"].value = "e"
        ws["H1"].value = "f"
        ws["F2"].value = "4"
        ws["G2"].value = "5"
        ws["H2"].value = "6"

        alignment = "horizontal"
        time_header_coord = ["F1", "F2"]
        ini = 6
        end = 8
        ti_iter = CleanMultipleColumns._time_index_iterator(ws, alignment,
                                                            time_header_coord,
                                                            ini, end)
        res = [i[0] for i in ti_iter]
        self.assertEqual(res, ["d 4", "e 5", "f 6"])


class CleanMultiColumnsMultiFreqTestCase(unittest.TestCase):

    # @unittest.skip("skip")

    def test_clean_time_index_case6(self):

        wb = load_workbook(
            os.path.join(abs_path("original"), "test_case6.xlsx"))
        ws = wb.active

        params = {"alignment": "horizontal",
                  "blank_rows": True,
                  "composed_headers": True,
                  "data_starts": 3,
                  "data_ends": 61,
                  "frequency": "AQQQQ",
                  "headers_coord": "B8-B28",
                  "continuity": False,
                  "missings": False,
                  "missing_value": None,
                  "multifrequency": True,
                  "series_names": None,
                  "time_composed": True,
                  "time_alignment": 0,
                  "time_multicolumn": True,
                  "time_format": str,
                  "time_header": True,
                  "time_header_coord": ["C4", "C6"]}

        CleanMultiColumnsMultiFreq().clean_time_index(ws, params)

        wb_exp = load_workbook(
            os.path.join(abs_path("expected"), "test_case6.xlsx"),
            data_only=True)

        # wb.save("test_case6_after_cleaning_index.xlsx")
        self.assertTrue(compare_cells(wb, wb_exp))

    def test_clean_time_index_case6_without_end(self):

        wb = load_workbook(
            os.path.join(abs_path("original"), "test_case6.xlsx"))
        ws = wb.active

        params = {"alignment": "horizontal",
                  "blank_rows": True,
                  "composed_headers": True,
                  "data_starts": 3,
                  "data_ends": None,
                  "frequency": "AQQQQ",
                  "headers_coord": "B8-B28",
                  "continuity": False,
                  "missings": False,
                  "missing_value": None,
                  "multifrequency": True,
                  "series_names": None,
                  "time_composed": True,
                  "time_alignment": 0,
                  "time_multicolumn": True,
                  "time_format": str,
                  "time_header": True,
                  "time_header_coord": ["C4", "C6"]}

        end = CleanMultiColumnsMultiFreq().clean_time_index(ws, params)

        wb_exp = load_workbook(
            os.path.join(abs_path("expected"), "test_case6.xlsx"),
            data_only=True)

        # wb.save("test_case6_after_cleaning_index.xlsx")
        self.assertTrue(compare_cells(wb, wb_exp))
        self.assertEqual(end, 61)


class BaseMultiFrequencyTestCase(unittest.TestCase):

    def test_init_progression_args(self):

        freq_args = BaseMultiFrequency._init_last_time_dict("AQQQQ")
        exp_freq_args = {"A": None, "Q": None}
        self.assertEqual(freq_args, exp_freq_args)

        freq_args = BaseMultiFrequency._init_last_time_dict("QQQQA")
        exp_freq_args = {"A": None, "Q": None}
        self.assertEqual(freq_args, exp_freq_args)

    def test_next_frequency(self):

        next_f, last_f = BaseMultiFrequency._next_frequency("AQQQQ", "AQ")
        self.assertEqual(next_f, "Q")

        next_f, last_f = BaseMultiFrequency._next_frequency("AQQQQ", "AQQQQ")
        self.assertEqual(next_f, "A")

        next_f, last_f = BaseMultiFrequency._next_frequency("QQQQA", "QQ")
        self.assertEqual(next_f, "Q")

        next_f, last_f = BaseMultiFrequency._next_frequency("QQQQA", "QQQQ")
        self.assertEqual(next_f, "A")


if __name__ == '__main__':
    nose.run(defaultTest=__name__)
