#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
strategies
----------------------------------

This module contains the high level strategies used by `xlseries` to parse
time data series inside excel files into Pandas DataFrames.
"""

import sys
import inspect
from pprint import pprint
import pandas as pd
import numpy as np
from openpyxl.cell import column_index_from_string

from parameters import Parameters
import clean_ti_strategies
import get_data_strategies


class BaseStrategy(object):

    """BaseStrategy class for all strategies."""

    def __init__(self, wb, input_params=Parameters()):
        self.wb = wb
        self.params = input_params

    # PUBLIC INTERFACE
    @classmethod
    def accepts(cls, wb):
        return cls._accepts(wb)

    def get_data_frames(self):
        return self._get_data_frames()


class ParameterDiscovery(BaseStrategy):

    """Strategy that aims to discover key parsing parameters."""

    # PRIVATE INTERFACE METHODS
    @classmethod
    def _accepts(cls, wb):
        return True

    def _get_data_frames(self):
        """Extract time data series and return them as data frames."""

        ws = self.wb.active

        # First: discover the parameters of the file
        self._discover_parameters(ws)

        # Second: clean the data
        self._clean_data(ws)

        # Third: get the data
        return self._get_data(ws)

    # HIGH LEVEL TASKS
    def _discover_parameters(self, ws):
        """Discover the parameters of the worksheet."""
        pass

    def _clean_data(self, ws):
        """Ensure data is clean to be processed with the parameters."""

        # 1. Clean time index
        for i_series in xrange(len(self.params.time_header_coord)):
            status_index = self._clean_time_index(ws, self.params[i_series])

        # 2. Clean data values
        for i in xrange(len(self.params.headers_coord)):
            status_values = self._clean_values(ws)

        return {"index": status_index, "values": status_values}

    def _get_data(self, ws):
        """Parse data using parameters and return it in data frames."""

        # 1. Build frames dict based on amount of frequencies
        frames_input_dict = {}
        for freq in self.params.frequency:
            frames_input_dict[freq] = {"columns": [], "data": []}

        # 2. Get name and values of each data series
        for i_series in xrange(len(self.params.headers_coord)):

            # iterate strategies looking for someone that accepts it
            params = self.params[i_series]
            name, values = None, None
            for strategy in get_data_strategies.get_strategies():
                if strategy.accepts(ws, params):
                    strategy_obj = strategy()
                    name, values = strategy_obj.get_data(ws, params)
                    break

            # raise exception if no strategy accepts the input
            if not name or not values:
                msg = "There is no strategy to deal with " + str(params)
                raise Exception(msg)

            frames_input_dict[params["frequency"]]["columns"].append(name)
            frames_input_dict[params["frequency"]]["data"].append(values)

        # 3. Build data frames
        dfs = []
        for period_range in self._get_period_ranges(ws):
            columns = frames_input_dict[period_range.freqstr]["columns"]

            data = frames_input_dict[period_range.freqstr]["data"]
            np_data = np.array(data).transpose()

            df = pd.DataFrame(index=period_range,
                              columns=columns,
                              data=np_data)

            dfs.append(df)

        return dfs

    # 2. CLEAN DATA methods
    @classmethod
    def _clean_time_index(cls, ws, params):

        for strategy in clean_ti_strategies.get_strategies():

            if strategy.accepts(ws, params):
                strategy_obj = strategy()
                status_index = strategy_obj.clean_time_index(ws, params)
                break

        return status_index

    @classmethod
    def _clean_values(cls, ws):
        status_data = True

        return status_data

    # 3. GET DATA methods
    def _get_period_ranges(self, ws):

        period_ranges = []

        for freq, ini_row, header_coord, end_row, time_alignement in \
            zip(self.params.frequency, self.params.data_starts,
                self.params.time_header_coord, self.params.data_ends,
                self.params.time_alignment):

            pr = self._get_period_range(ws, freq, ini_row, header_coord,
                                        end_row, time_alignement)
            period_ranges.append(pr)

        return period_ranges

    def _get_period_range(self, ws, freq, ini_row, header_coord, end_row,
                          time_alignement):
        col = column_index_from_string(ws[header_coord].column)
        period_range = pd.period_range(ws.cell(row=ini_row + time_alignement,
                                               column=col).value,
                                       ws.cell(row=end_row + time_alignement,
                                               column=col).value,
                                       freq=freq)

        return period_range


def get_strategies_names():
    """Returns a list of the parsers names, whith no Base classes."""

    list_cls_tuple = inspect.getmembers(sys.modules[__name__], inspect.isclass)
    list_cls_names = [cls_tuple[0] for cls_tuple in list_cls_tuple]
    list_no_base_cls_names = [cls_name for cls_name in list_cls_names
                              if cls_name[:4] != "Base" and
                              cls_name != "Parameters"]

    return list_no_base_cls_names


def get_strategies():
    """Returns a list of references to the parsers classes."""

    return [globals()[cls_name] for cls_name in get_strategies_names()]


if __name__ == '__main__':
    pprint(sorted(get_strategies_names()))
