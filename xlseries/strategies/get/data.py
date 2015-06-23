#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
data

This module contains strategies to get data from a spreadsheet.
"""

from __future__ import unicode_literals
import sys
import inspect
from pprint import pprint
import arrow
import datetime
import numpy as np
from openpyxl.cell import column_index_from_string
from openpyxl.cell import get_column_letter
from unidecode import unidecode

import xlseries.utils.strategies_helpers
from xlseries.utils.time_manipulation import increment_time


class BaseGetDataStrategy(object):

    """Implements the interface for all get data strategies."""

    # PUBLIC INTERFACE
    @classmethod
    def accepts(cls, ws, params):
        return cls._accepts(ws, params)

    def get_data(self, ws, params):
        return self._get_data(ws, params)

    # PRIVATE
    @classmethod
    def _accepts(cls, ws, params):
        """Accepts inputs that are accepted by all the base classes."""
        return cls._base_cond(ws, params)

    def _get_data(self, ws, params):
        name = self._get_name(ws, params["headers_coord"])
        # print name
        values_list = self._get_values(ws, params)
        # print values_list

        return [(name, values) for values in values_list]

    @classmethod
    def _get_name(cls, ws, header_coord):
        return unidecode(ws[header_coord].value).strip()

    def _get_values(self, ws, params):

        # create iterator of values
        iter_values = self._values_iterator(ws, params["alignment"],
                                            params["headers_coord"],
                                            params["data_starts"],
                                            params["data_ends"])

        values_dict = {}
        for value, index in iter_values:
            # print value, index
            new_value = self._handle_new_value(values_dict.values(), value,
                                               params["missings"],
                                               params["missing_value"],
                                               params["blank_rows"])

            if self._value_to_be_added(new_value, index, ws, params):
                frequency = self._get_frequency(params["frequency"])
                if frequency not in values_dict:
                    values_dict[frequency] = []
                values_dict[frequency].append(new_value)

        # fill the missing values if they are implicit
        if (params["missings"] and params["missing_value"] == "Implicit" and
                not params["multifrequency"]):
            a = values_dict[frequency]
            a = self._fill_implicit_missings(ws,
                                             values_dict[frequency],
                                             params["frequency"],
                                             params["time_header_coord"],
                                             params["data_starts"],
                                             params["data_ends"])
            values_dict[frequency] = a

        return values_dict.values()

    @classmethod
    def _base_cond(cls, ws, params):
        """Check that all base classes accept the input."""
        for base in cls.__bases__:
            if (base is not BaseGetDataStrategy and
                    not base._accepts(ws, params)):
                return False
        return True

    @classmethod
    def _values_iterator(cls, ws, alignment, header_coord, ini, end):

        if alignment == "vertical":
            col = ws[header_coord].column
            for row in xrange(ini, end + 1):
                yield (ws.cell(coordinate=col + unicode(row)).value, row)

        elif alignment == "horizontal":
            row = ws[header_coord].row
            for col in xrange(ini, end + 1):
                yield (ws.cell(column=col, row=row).value, col)

        else:
            raise Exception("Series alignment must be 'vertical' or " +
                            "'horizontal', not " + repr(alignment))

    @classmethod
    def _valid_value(cls, value):
        """Check if a value is likely to be a series data value."""

        RV = True

        try:
            float(value)
        except:
            RV = False

        return RV


class BaseSingleFrequency():

    @classmethod
    def _accepts(cls, ws, params):
        return not params["multifrequency"]

    # PRIVATE
    @classmethod
    def _fill_implicit_missings(cls, ws, values, frequency, time_header_coord,
                                ini_row, end_row):
        """Fill time holes in the series with missing data."""

        col = ws[time_header_coord].column

        new_values = []
        ini_time_value = arrow.get(ws.cell(coordinate=col +
                                           unicode(ini_row)).value)
        exp_time_value = ini_time_value
        for row, (i_value, value) in zip(xrange(ini_row, end_row + 1),
                                         enumerate(values)):
            obs_time_value = arrow.get(
                ws.cell(coordinate=col + unicode(row)).value)

            # fill time holes in the series with missing data
            while exp_time_value < obs_time_value:
                new_values.append(np.nan)
                exp_time_value = increment_time(exp_time_value, 1, frequency)

            new_values.append(values[i_value])
            exp_time_value = increment_time(exp_time_value, 1, frequency)

        return new_values

    @classmethod
    def _get_frequency(cls, frequency):
        return frequency


class BaseMultiFrequency():

    def __init__(self):
        self.last_frequency = None

    @classmethod
    def _accepts(cls, ws, params):
        return params["multifrequency"]

    def _get_frequency(self, frequency):
        freq, self.last_frequency = self._next_frequency(frequency,
                                                         self.last_frequency)
        return freq

    @classmethod
    def _next_frequency(cls, frequency, last_frequency=None):
        """Calculates what frequency go next.

        In single frequency series, returns the frequency parameter. This
        method gains relevance in multifrequency series, where a tracking of
        the last frequency is needed to know what frequency should be applied
        for a time value in a certain point of the time index.
        """

        if not len(frequency) > 1:
            return frequency

        if not last_frequency or last_frequency == frequency:
            freq = frequency[0]
            last_frequency = freq
        else:
            # print frequency.partition(last_frequency)
            freq = frequency.partition(last_frequency)[2][0]
            assert len(freq) == 1, "Freq must have only one character."
            last_frequency += freq

        # print freq, last_frequency

        return freq, last_frequency


class BaseContinuous():

    """Get data from continuous series."""

    @classmethod
    def _accepts(cls, ws, params):
        return params["continuity"]

    @classmethod
    def _value_to_be_added(cls, value, index, ws, params):
        """Check if a value should be added or not."""
        return value is not None

    @classmethod
    def _handle_new_value(cls, values, value, missings, missing_value,
                          blank_rows):

        if blank_rows and value is None:
            return None

        if missings:
            if value != missing_value:
                args_without_values = locals()
                del args_without_values["values"]
                try:
                    return float(value)
                except:
                    print args_without_values
                    raise Exception("Value is not valid " + unicode(value))
            else:
                return np.nan
        else:
            return float(value)


class BaseNonContinuous():

    """Get data from non continuous series."""

    @classmethod
    def _accepts(cls, ws, params):
        return not params["continuity"]

    @classmethod
    def _value_to_be_added(cls, value, index, ws, params):
        """Check if a value should be added or not.

        Value shouldn't be None and the row should correspond to a valid time
        value in the time index."""

        if params["alignment"] == "vertical":
            time_col = ws[params["time_header_coord"]].column
            time_coord = time_col + unicode(index + params["time_alignment"])
            time_value = ws[time_coord].value

        elif params["alignment"] == "horizontal":
            time_row = ws[params["time_header_coord"]].row
            time_value = ws.cell(column=index + params["time_alignment"],
                                 row=time_row).value

        else:
            raise Exception("Series alignment must be 'vertical' or " +
                            "'horizontal', not " + repr(params["alignment"]))

        return value is not None and type(time_value) == datetime.datetime

    @classmethod
    def _handle_new_value(cls, values, value, missings, missing_value,
                            blank_rows):

        new_value = None
        if missings:
            if value == missing_value:
                new_value = np.nan
            elif cls._valid_value(value):
                new_value = float(value)
            # values that are not valid nor missings
            else:
                pass
        else:
            if cls._valid_value(value):
                new_value = float(value)

        return new_value


class GetSingleFrequencyContinuous(BaseSingleFrequency, BaseContinuous,
                                   BaseGetDataStrategy):

    """Get data with a single frequency and continous layout."""
    pass


class GetSingleFrequencyNonContinuous(BaseSingleFrequency, BaseNonContinuous,
                                      BaseGetDataStrategy):

    """Get data with a single frequency and data layout interrupted.

    The interruption is due to strange strings or values that should not be
    taken into account when gathering values. Series interrupted only by blank
    rows do not need this strategy."""
    pass


class GetMultiFrequencyContinuous(BaseMultiFrequency, BaseContinuous,
                                  BaseGetDataStrategy):

    """Get data from a multifrequency series and continous layout."""
    pass


class GetMultiFrequencyNonContinuous(BaseMultiFrequency, BaseNonContinuous,
                                     BaseGetDataStrategy):

    """Get data from a multifrequency series and data layout interrupted.

    The interruption is due to strange strings or values that should not be
    taken into account when gathering values. Series interrupted only by blank
    rows do not need this strategy."""
    pass


def get_strategies():
    return xlseries.utils.strategies_helpers.get_strategies()

if __name__ == '__main__':
    pprint(sorted(xlseries.utils.strategies_helpers.get_strategies_names()))
