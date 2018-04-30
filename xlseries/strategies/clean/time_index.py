#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
clean_ti_strategies

This module contains strategies to parse and clean time indexes in worksheets
cotaining time data series.

Warning! Do not import other classes directly "from module import Class",
except if they are custom exceptions.
Rather import the module in which the Class is defined and use it like
"module.Class". All the classes defined in this modul namespace are
automatically taken by "get_strategies" and exposed to the user.
"""

import arrow
from pprint import pprint
from pprint import pformat
from openpyxl.utils import get_column_letter, column_index_from_string
import datetime

from xlseries.strategies.clean.parse_time import DayOutOfRange, MonthOutOfRange
from xlseries.strategies.clean.parse_time import NoTimeValue
from xlseries.strategies.clean.parse_time import NoPossibleTimeValue
import xlseries.utils.strategies_helpers
from xlseries.utils.time_manipulation import increment_time
import xlseries.strategies.clean.parse_time as parse_time_strategies


# CUSTOM EXCEPTIONS
class BaseProgressionError(ValueError):

    """Raised when the progression of a time value is wrong."""

    def get_msg(self, curr_time, exp_time, last_time):
        return " ".join(["Last:", str(last_time),
                         "\nExpected:", str(exp_time),
                         "\nCurrent:", str(curr_time)])


class TimeValueGoingBackwards(BaseProgressionError):

    """Raised when a time value is going backwards.

    The parser observe that a time value is going backwards and
    existent strategies can't deal with it."""

    def __init__(self, curr_time, exp_time, last_time):
        msg = self.get_msg(curr_time, exp_time, last_time)
        super(TimeValueGoingBackwards, self).__init__(msg)


class TimeValueGoingForth(BaseProgressionError):

    """Raised when time value is going forth, when not supposed to.

    The parser observe that a time value is going forth than expected
    and existent strategies can't deal with it."""

    def __init__(self, curr_time, exp_time, last_time):
        msg = self.get_msg(curr_time, exp_time, last_time)
        super(TimeValueGoingForth, self).__init__(msg)


class ParseTimeImplementationError(NotImplementedError):

    """Raised when parsing to date data structure is impossible.

    There is no strategy to deal with the time value that is trying to be
    parsed."""

    def __init__(self, curr_time, last_time, next_time, params):
        msg = " ".join(["No strategy to parse time.",
                        "\nCurrent:", repr(
                            curr_time), repr(type(curr_time)),
                        "\nLast:", repr(last_time), repr(type(last_time)),
                        "\nNext:", repr(next_time), repr(type(next_time)),
                        "\nParameters: ", pformat(params)])
        super(ParseTimeImplementationError, self).__init__(msg)


class SameTimeValue(ValueError):

    """Raised if the value is the same as the last one."""

    def __init__(self, value, last_value):
        msg = " ".join(["Current value", str(value),
                        "is the same as the last value", str(last_value)])
        super(SameTimeValue, self).__init__(msg)


# STRATEGIES
class BaseCleanTiStrategy(object):

    """BaseCleanTiStrategy class for all time index cleaning strategies."""

    NO_TIME_VALUE_LIMIT = 40

    def __init__(self, time_parser=None):
        self.time_parser = time_parser

    # PUBLIC INTERFACE
    @classmethod
    def accepts(cls, ws, params):
        return cls._accepts(ws, params)

    def clean_time_index(self, ws, params):
        return self._clean_time_index(ws, params)

    # PRIVATE main methods
    @classmethod
    def _accepts(cls, ws, params):
        return cls._base_cond(ws, params)

    def _clean_time_index(self, ws, params):
        """Parse time strings into time values, cleaning the time index.

        If a value in a cell should be a time value, replace it with the clean
        time value."""
        # import pdb; pdb.set_trace()
        p = params
        # create iterator of time index values
        iter_time_index = self._time_index_iterator(
            ws, p["alignment"], p["time_header_coord"], p["data_starts"],
            p["data_ends"])

        last_time = None
        no_time_value_count = 0
        for curr_time, next_time, write_time_cell in iter_time_index:

            # only clean if the value is expected to be a time value
            if self._must_be_time_value(curr_time, next_time, last_time):
                no_time_value_count = 0

                try:
                    curr_time = self._parse_time(params, curr_time, last_time,
                                                 next_time)

                    # correct typos checking for a healthy time progression
                    curr_time = self._correct_progression(last_time,
                                                          curr_time,
                                                          p["frequency"],
                                                          p["missings"],
                                                          p["missing_value"])

                    # avoid writing the same time value again, except in the
                    # multifrequency case, where year could be equal to the
                    # first quarter... TODO: better treatment for multifreq
                    if curr_time == last_time and len(p["frequency"]) == 1:
                        raise SameTimeValue(curr_time, last_time)

                    # write the clean value to the spreadsheet
                    write_time_cell.value = curr_time.datetime
                    last_time = curr_time

                # this is the only case that _must_be_time_value is not
                # expected to avoid before calling _parse_time, it's a mistake
                # of the excel designers in the time index
                except (DayOutOfRange, MonthOutOfRange):
                    write_time_cell.value = None

                except (ParseTimeImplementationError, NoPossibleTimeValue,
                        NoTimeValue, SameTimeValue, AssertionError):

                    if not p["data_ends"]:
                        return self._estimate_end(p["alignment"],
                                                  write_time_cell,
                                                  p["data_starts"],
                                                  p["time_alignment"])
                    else:
                        raise

                except:
                    raise

            elif (no_time_value_count < self.NO_TIME_VALUE_LIMIT and
                  (not p["continuity"] or p["blank_rows"])):
                no_time_value_count += 1

            else:
                break

        return self._estimate_end(p["alignment"], write_time_cell,
                                  p["data_starts"], p["time_alignment"])

    @classmethod
    def _must_be_time_value(cls, value, next_time, last_time):
        return ((value is not None) and (len(str(value).strip()) > 0))

    @classmethod
    def _estimate_end(cls, alignment, last_cell, start, time_alignment):
        if alignment == "vertical":
            while (not isinstance(last_cell.value, datetime.datetime) and
                    last_cell.row > start):
                last_cell = last_cell.offset(row=-1)

            end = last_cell.row - time_alignment
            msg = "End must be greater than start! End: {} / Start: {}".format(
                repr(end).ljust(6), start
            )
            assert end and end > start, msg
            return end

        else:
            while (not isinstance(last_cell.value, datetime.datetime) and
                    last_cell.column > start):
                # print type(last_cell.value), last_cell.value, last_cell.row,
                # last_cell.column, type(last_cell.offset(row=2).value),
                # last_cell.offset(row=2).value, last_cell.offset(row=2).row,
                # last_cell.offset(row=2).column
                last_cell = last_cell.offset(column=-1)

            end = column_index_from_string(last_cell.column) - time_alignment
            msg = "End must be greater than start! End: {} | Start: {}".format(
                end, start
            )
            assert end and end > start, msg
            return end

    # PRIVATE time index iterator methods
    @classmethod
    def _time_index_iterator(cls, ws, alignment, time_header_coord, ini,
                             end=None):

        if alignment == "vertical":
            end = end or cls._get_row_boundary(ws, time_header_coord, ini)
            for row in range(ini, end + 1):
                curr_time = cls._get_time_value(ws, time_header_coord,
                                                f_row=row)
                next_time = cls._get_time_value(ws, time_header_coord,
                                                f_row=row + 1)
                col = cls._time_header_cell(ws, time_header_coord).column
                write_time_cell = ws[col + str(row)]

                yield (curr_time, next_time, write_time_cell)

        elif alignment == "horizontal":
            end = end or cls._get_column_boundary(ws, time_header_coord, ini)
            for col in range(ini, end + 1):
                curr_time = cls._get_time_value(ws, time_header_coord,
                                                f_col=get_column_letter(col))
                next_time = cls._get_time_value(
                    ws, time_header_coord, f_col=get_column_letter(col + 1))
                row = cls._time_header_cell(ws, time_header_coord).row
                write_time_cell = ws.cell(column=col, row=row)

                yield (curr_time, next_time, write_time_cell)

        else:
            raise Exception("Series alignment must be 'vertical' or " +
                            "'horizontal', not " + repr(alignment))

    @classmethod
    def _get_row_boundary(cls, ws, time_header_coord, ini):
        """Returns the pressumed last row of a column."""
        raise NotImplementedError("Getting the row boundary must be " +
                                  "implemented in a subclass.")

    @classmethod
    def _get_time_value(cls, ws, time_header_coord, f_row=None, f_col=None):
        """Returns the time value corresponding a certain series and row."""
        raise NotImplementedError("Getting the time value must be " +
                                  "implemented in a subclass.")

    @classmethod
    def _time_header_cell(cls, ws, time_header_coord):
        """Returns the column where clean time index shouls be written."""
        if isinstance(time_header_coord, list):
            return ws[time_header_coord[0]]
        else:
            return ws[time_header_coord]

    # PRIVATE methods to parse time values
    def _parse_time(self, params, curr_time, last_time=None, next_time=None):
        """Try to parse any value into a proper date format.

        Iterate a pool of strategies looking one that understands the format of
        the time value to be parsed and declares to be able to parse it into
        a date format.

        Args:
            params: Parameters of the series being analyzed.
            curr_time: Value to be parsed into a date format.
            last_time: Last value parsed into time value.
            next_time: Next value to be parsed into time value.

        Returns:
            An arrow.Arrow object expressing a date.
        """
        msg = "parse_time strategies must assure a valid time value!"
        # first, try to use the parser used last time
        if self.time_parser:
            try:
                time_value = self.time_parser.parse_time(params, curr_time,
                                                         last_time, next_time)
                assert isinstance(time_value, arrow.Arrow), msg

                return time_value

            except (DayOutOfRange, MonthOutOfRange) as inst:
                raise inst

            except:
                pass
        # import pdb; pdb.set_trace()
        # if last parser doesn't work (or there is None), search again
        for strategy in parse_time_strategies.get_strategies():
            if strategy.accepts(params, curr_time, last_time, next_time):
                self.time_parser = strategy()
                time_value = self.time_parser.parse_time(params, curr_time,
                                                         last_time, next_time)

                assert isinstance(time_value, arrow.Arrow), msg

                return time_value

        raise ParseTimeImplementationError(curr_time, last_time, next_time,
                                           params)

    # PRIVATE methods to correct progression
    @classmethod
    def _correct_progression(cls, last_time, curr_time,
                             freq, missings, missing_value=None):

        # without a last_time the progression cannot be corrected
        if not last_time:
            return curr_time

        exp_time = increment_time(last_time, 1, freq)
        assert isinstance(exp_time, arrow.Arrow)
        assert isinstance(last_time, arrow.Arrow) or not last_time
        assert isinstance(curr_time, arrow.Arrow)

        # everything is ok!
        if exp_time == curr_time:
            return curr_time

        # going back
        if curr_time < last_time:
            if cls._time_value_typo(curr_time, exp_time):
                return exp_time
            else:
                raise TimeValueGoingBackwards(curr_time, exp_time,
                                              last_time)

        # going forth with no missings allowed
        going_forth = curr_time > last_time
        if going_forth and not missings:
            if cls._time_value_typo(curr_time, exp_time):
                return exp_time
            else:
                raise TimeValueGoingForth(curr_time, exp_time,
                                          last_time)

        # going forth with implicit missings
        max_forth_time_value = increment_time(last_time,
                                              cls._max_forth_units(freq), freq)
        going_too_forth = curr_time > max_forth_time_value
        if going_too_forth and missings and missing_value == "Implicit":
            forth_time_value = cls._forth_time_value_typo(curr_time,
                                                          max_forth_time_value)
            if forth_time_value:
                return forth_time_value
            else:
                return False

        # everything should be ok
        else:
            return curr_time

    @classmethod
    def _max_forth_units(cls, freq):
        max_forth_units = {"D": 20,
                           "M": 2,
                           "Q": 1,
                           "A": 1}
        if freq in max_forth_units:
            return max_forth_units[freq]
        else:
            return max(max_forth_units.values())

    @classmethod
    def _time_value_typo(cls, curr_time, exp_time):
        """Check if a time value could have a human typo.

        It relies in the idea that if the current time value being
        parsed is equal in two of the three parameters of a date
        (day, month and year), the different one is a human typo.

        Args:
            curr_time: Time value to be analyzed for a typo.
            exp_time: Expected time value to compare.
        """

        matches = [(curr_time.day == exp_time.day),
                   (curr_time.month == exp_time.month),
                   (curr_time.year == exp_time.year)]

        if matches.count(True) == 2:
            return True
        else:
            return False

    @classmethod
    def _forth_time_value_typo(cls, curr_time, max_forth_time_value):
        """Check for a typo in day, month or year.

        This check is based on the idea that one of the three elements of
        the date is wrong because of a typo that makes the date greater than a
        possible maximum. When one of the possibilities generates a date lower
        than maximum it means the typo was succesfully corrected.

        Args:
            curr_time (arrow.Arrow): Time value being analyzed because
                it is greater than it should be.
            max_forth_time_value (arrow.Arrow): Maximum possible time value for
                curr_time being analyzed.

        Returns:
            arrow.Arrow or None: A fixed time value removing the typo or None
                if the typo couldn't be fixed.
        """

        day_typo = arrow.get(curr_time.year,
                             curr_time.month,
                             max_forth_time_value.day)

        month_typo = arrow.get(curr_time.year,
                               max_forth_time_value.month,
                               curr_time.day)

        year_typo = arrow.get(max_forth_time_value.year,
                              curr_time.month,
                              curr_time.day)

        for possible_typo in [day_typo, month_typo, year_typo]:
            if possible_typo < max_forth_time_value:
                return possible_typo

        return None


class BaseAccepts():

    """Provide the basic accepts conditions resolution."""

    @classmethod
    def _accepts(cls, ws, params):
        return cls._base_cond(ws, params)

    @classmethod
    def _base_cond(cls, ws, params):
        """Check that all base classes accept the input."""
        for base in cls.__bases__:
            if (
                base is not BaseCleanTiStrategy and
                base is not cls and
                (hasattr(base, "_accepts") and not base._accepts(ws, params))
            ):
                return False
        return True


class BaseSingleTable():

    """Presumes the sheet has a single table on it."""

    # PRIVATE INTERFACE METHODS
    @classmethod
    def _accepts(cls, ws, params):
        return True

    @classmethod
    def _get_row_boundary(cls, ws, time_header_coord, ini):
        """Returns the pressumed last row of a column."""
        return ws.max_row

    @classmethod
    def _get_column_boundary(cls, ws, time_header_coord, ini):
        """Returns the pressumed last column of a row."""
        return ws.max_column


class BaseMultiTable():

    """Presumes the sheet has many tables on it."""

    # PRIVATE INTERFACE METHODS
    @classmethod
    def _accepts(cls, ws, params):
        return (
            not params["blank_rows"] and
            params["continuity"] and not
            params["data_ends"]
        )

    @classmethod
    def _get_row_boundary(cls, ws, time_header_coord, ini):
        """Returns the last non empty row of a table, not the worksheet."""
        i = 0
        while ws[time_header_coord].offset(row=i).value:
            i += 1
        return ws[time_header_coord].offset(row=i).row

    @classmethod
    def _get_column_boundary(cls, ws, time_header_coord, ini):
        """Returns the last non empty column of a table, not the worksheet."""
        i = 0
        while ws[time_header_coord].offset(column=i).value:
            i += 1
        return ws[time_header_coord].offset(column=i).column


class BaseSingleColumn():

    """Clean time indexes that use a single column."""

    # PRIVATE INTERFACE METHODS
    @classmethod
    def _accepts(cls, ws, params):
        return not params["time_multicolumn"]

    @classmethod
    def _get_time_value(cls, ws, time_header_coord, f_row=None, f_col=None):
        """Returns the time value corresponding a certain series and row."""
        assert not isinstance(
            time_header_coord, list), "Time header should be a str."

        col = str(f_col or ws[time_header_coord].column)
        row = str(f_row or ws[time_header_coord].row)

        return ws[col + row].value


class BaseMultipleColumns():

    """Clean time indexes that use a single column."""

    # PRIVATE INTERFACE METHODS
    @classmethod
    def _accepts(cls, ws, params):
        return params["time_multicolumn"]

    @classmethod
    def _get_time_value(cls, ws, time_header_coord, f_row=None, f_col=None):
        """Returns the time value corresponding a certain series and row.

        Concatenate all the values of the time header columns in a unique
        string."""
        assert isinstance(time_header_coord,
                          list), "Time header should be a list."

        time_value_list = []

        # print(time_header_coord)
        for coord in time_header_coord:
            col = str(f_col or ws[coord].column)
            row = str(f_row or ws[coord].row)
            value = ws[col + row].value

            msg = "there shouldn't be time values in multicolumn!"
            assert not isinstance(value, datetime.datetime), msg

            if value and value != "None":
                time_value_list.append(cls._safe_str(value))

        time_value = " ".join(time_value_list)

        if len(time_value.strip()) > 0:
            return time_value
        else:
            return None

    @classmethod
    def _safe_str(cls, value):
        """Check if the value is a number before make it unicode."""
        try:
            return str(int(value))
        except:
            return str(value)


class BaseOffsetTi():

    """Clean time indexes where time alignment is offset, sharing the same
    column with the data."""

    @classmethod
    def _accepts(cls, ws, params):
        return params["time_alignment"] != 0

    @classmethod
    def _must_be_time_value(cls, value, next_time, last_time):
        base_cond = BaseCleanTiStrategy._must_be_time_value(value,
                                                            next_time,
                                                            last_time)
        return base_cond and not isinstance(value, float)


class BaseNoOffsetTi():

    """Clean time indexes where time alignment is 0, the most common case."""

    @classmethod
    def _accepts(cls, ws, params):
        return params["time_alignment"] == 0


class BaseSingleFrequency():

    """Only accepts single frequency series."""

    @classmethod
    def _accepts(cls, ws, params):
        return len(params["frequency"]) == 1


class BaseMultiFrequency():

    """Reimplements private methods for multifrequency series."""

    def __init__(self, *args, **kwargs):
        BaseCleanTiStrategy.__init__(self, *args, **kwargs)
        self.last_time = {}
        self.last_frequency = None

    @classmethod
    def _accepts(cls, ws, params):
        return len(params["frequency"]) > 1

    def _correct_progression(self, last_time, curr_time,
                             frequency, missings, missing_value=None):
        if len(self.last_time) == 0:
            self.last_time = self._init_last_time_dict(frequency)

        # frequency and last_time are replaced simulating two single frequency
        # series instead of one multifrequency
        freq, self.last_frequency = self._next_frequency(frequency,
                                                         self.last_frequency)
        last_time = self.last_time[freq]

        superclass = BaseCleanTiStrategy
        curr_time = superclass._correct_progression(last_time, curr_time,
                                                    freq, missings,
                                                    missing_value)
        self.last_time[freq] = curr_time
        return curr_time

    @classmethod
    def _init_last_time_dict(cls, frequency):
        """Create a dictionary for _correct_progression arguments.

        Each entry is a different frequency, in a multifrequency series."""
        freqs = "".join(set(frequency))
        return {freq: None for freq in freqs}

    @classmethod
    def _next_frequency(cls, frequency, last_frequency=None):
        """Calculates what frequency go next.

        In single frequency series, returns the frequency parameter. This
        method gains relevance in multifrequency series, where a tracking of
        the last frequency is needed to know what frequency should be applied
        for a time value in a certain point of the time index.
        """

        if not len(frequency) > 1:
            return frequency

        if not last_frequency or last_frequency == frequency:
            freq = frequency[0]
            last_frequency = freq
        else:
            freq = frequency.partition(last_frequency)[2][0]
            assert len(freq) == 1, "Freq must have only one character."
            last_frequency += freq

        return freq, last_frequency


def get_strategies():
    custom = xlseries.utils.strategies_helpers.get_strategies()

    combinations = []
    for table in [BaseSingleTable, BaseMultiTable]:
        for offset in [BaseNoOffsetTi, BaseOffsetTi]:
            for freq in [BaseSingleFrequency, BaseMultiFrequency]:
                for col in [BaseSingleColumn, BaseMultipleColumns]:

                    name = "{}{}{}{}".format(
                        table.__name__,
                        col.__name__,
                        freq.__name__,
                        offset.__name__
                    )
                    bases = (BaseAccepts, table, col, freq,
                             offset, BaseCleanTiStrategy)
                    parser = type(name, bases, {})

                    combinations.append(parser)

    return custom + combinations


if __name__ == '__main__':
    pprint(sorted(xlseries.utils.strategies_helpers.get_strategies_names()))
