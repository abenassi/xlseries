#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
test_get_data_strategies
----------------------------------

Tests for `get_data_strategies` module.
"""

import unittest
import nose
from openpyxl import load_workbook

from xlseries.utils import get_data_frames, approx_equal
from xlseries.parameters import Parameters
from xlseries.get_data_strategies import GetSingleFrequencyData
from xlseries.clean_ti_strategies import CleanSimpleTi
from utils import compare_list_values


class MissingsTestCase(unittest.TestCase):

    def _get_values(self, ws, ini_row, end_row, col):

        values = []
        i_row = ini_row
        while i_row <= end_row:
            values.append(ws.cell(row=i_row, column=col).value)
            i_row += 1

        return values

    def test_fill_implicit_missings(self):
        test_wb = load_workbook("cases/test_case2.xlsx")
        params = Parameters("cases/test_case2_params.json")
        strategy = GetSingleFrequencyData

        ws = test_wb.active

        ini_row = 5
        end_row = 2993
        col = 4

        values = self._get_values(ws, ini_row, end_row, col)
        frequency = "D"
        time_header_coord = "C4"

        CleanSimpleTi.clean_time_index(ws, params[0])

        new_values = strategy._fill_implicit_missings(ws, values, frequency,
                                                      time_header_coord,
                                                      ini_row,
                                                      end_row)

        exp_dfs = get_data_frames("cases/test_case2_exp.xlsx")
        exp_values = [value[0] for value in exp_dfs[0].values]

        self.assertEqual(len(new_values), len(exp_values))

        with open("record.txt", "wb") as f:
            for value1, value2 in zip(new_values, exp_values):
                f.write(str(value1).ljust(20) + str(value2).ljust(20) +
                        str(approx_equal(value1, value2, 0.001)).ljust(20) +
                        "\n")

        self.assertTrue(compare_list_values(new_values, exp_values))


if __name__ == '__main__':
    nose.run(defaultTest=__name__)
