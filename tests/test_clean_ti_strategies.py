#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
test_clean_ti_strategies
----------------------------------

Tests for `clean_ti_strategies` module.
"""

import unittest
import nose
import datetime
from openpyxl import load_workbook

from xlseries.clean_ti_strategies import CleanSimpleTi
from xlseries.utils import compare_cells


# @unittest.skip("skip")
class CleanSimpleTiTestCase(unittest.TestCase):

    def setUp(self):
        self.strategy = CleanSimpleTi

    def test_correct_progression(self):

        last_time_value = datetime.datetime(2011, 7, 5)
        curr_time_value = datetime.datetime(2011, 5, 6)
        freq = "D"
        missings = True
        missing_value = "Implicit"

        new_time_value = self.strategy._correct_progression(last_time_value,
                                                            curr_time_value,
                                                            freq, missings,
                                                            missing_value)
        exp_time_value = datetime.datetime(2011, 7, 6)

        self.assertEqual(new_time_value, exp_time_value)

    def test_parse_time(self):

        value = "17-12.09"
        time_format = datetime.datetime

        new_time_value = self.strategy._parse_time(value, time_format)
        exp_time_value = datetime.datetime(2009, 12, 17)

        self.assertEqual(new_time_value, exp_time_value)

    def test_clean_time_index(self):

        wb = load_workbook("cases/test_case2.xlsx")
        ws = wb.active

        clean_ci = {"time_alignment": 0,
                    "time_format": datetime.datetime,
                    "time_header_coord": "C4",
                    "ini_row": 5,
                    "end_row": 2993,
                    "frequency": "D",
                    "missings": True,
                    "missing_value": "Implicit"}

        self.strategy._clean_time_index(ws, clean_ci)

        wb_exp = load_workbook("cases/test_case2_clean_index.xlsx")

        wb.save("test_case2_after_cleaning_index.xlsx")
        self.assertTrue(compare_cells(wb, wb_exp))


if __name__ == '__main__':
    nose.run(defaultTest=__name__)
