#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
test_clean_ti_strategies

Tests for `clean_ti_strategies` module.
"""

import unittest
import nose
import arrow
import datetime
import os
from openpyxl import load_workbook

from xlseries.strategies.clean.time_index import CleanSingleColumnTi
from xlseries.strategies.clean.time_index import CleanMultipleColumnsTiConcat
from xlseries.strategies.clean.time_index import BaseCleanTiStrategy
from xlseries.strategies.clean.time_index import TimeValueGoingBackwards
from xlseries.strategies.clean.time_index import TimeValueGoingForth
from xlseries.utils.xl_methods import compare_cells
from xlseries.utils.case_loaders import load_parameters_case
from xlseries.utils.path_finders import abs_path


class BaseCleanTiStrategyTestCase(unittest.TestCase):

    def test_correct_progression_backwards_exception(self):
        last = arrow.get(2015, 5, 9)
        curr = arrow.get(2015, 2, 15)
        freq = "D"
        missings = False

        with self.assertRaises(TimeValueGoingBackwards):
            BaseCleanTiStrategy._correct_progression(last, curr, freq,
                                                     missings)

    def test_correct_progression_forth_exception(self):
        last = arrow.get(2015, 5, 9)
        curr = arrow.get(2016, 6, 30)
        freq = "M"
        missings = False

        with self.assertRaises(TimeValueGoingForth):
            BaseCleanTiStrategy._correct_progression(last, curr, freq,
                                                     missings)


# @unittest.skip("skip")
class CleanSingleColumnTiTestCase(unittest.TestCase):

    # @unittest.skip("skip")
    def test_correct_progression(self):

        # progression wrong because going to the past
        last_time_value = arrow.get(2011, 7, 5)
        curr_time_value = arrow.get(2011, 5, 6)
        freq = "D"
        missings = True
        missing_value = "Implicit"

        new_time_value = CleanSingleColumnTi._correct_progression(
            last_time_value,
            curr_time_value,
            freq, missings,
            missing_value)
        exp_time_value = arrow.get(2011, 7, 6)

        self.assertEqual(new_time_value, exp_time_value)

        # progression wrong because going to the future
        curr_time_value = arrow.get(2011, 8, 6)
        new_time_value = CleanSingleColumnTi._correct_progression(
            last_time_value,
            curr_time_value,
            freq, missings,
            missing_value)

        self.assertEqual(new_time_value, exp_time_value)

    # @unittest.skip("skip")
    def test_parse_time(self):

        value = "17-12.09"
        last_time = arrow.get(2009, 12, 16)

        params = load_parameters_case(2)
        # print repr(params[0])

        new_time_value = CleanSingleColumnTi._parse_time(params[0], value,
                                                         last_time)

        exp_time_value = arrow.get(2009, 12, 17)

        self.assertEqual(new_time_value, exp_time_value)

    # @unittest.skip("skip")
    def test_clean_time_index_case2(self):

        wb = load_workbook(
            os.path.join(abs_path("original"), "test_case2.xlsx"))
        ws = wb.active

        params = {"time_alignment": 0,
                  "time_format": datetime.datetime,
                  "continuity": True,
                  "blank_rows": True,
                  "time_header_coord": "C4",
                  "data_starts": 5,
                  "data_ends": 2993,
                  "frequency": "D",
                  "missings": True,
                  "missing_value": "Implicit",
                  "time_multicolumn": False,
                  "time_composed": False}

        CleanSingleColumnTi._clean_time_index(ws, params)

        wb_exp = load_workbook(
            os.path.join(abs_path("expected"), "test_case2.xlsx"))

        # wb.save("test_case2_after_cleaning_index.xlsx")
        self.assertTrue(compare_cells(wb, wb_exp))

    # @unittest.skip("skip")
    def test_clean_time_index_case5(self):

        wb = load_workbook(
            os.path.join(abs_path("original"), "test_case5.xlsx"))
        ws = wb.active

        params = {"time_alignment": 0,
                  "time_format": str,
                  "continuity": False,
                  "blank_rows": True,
                  "time_header_coord": "A18",
                  "data_starts": 28,
                  "data_ends": 993,
                  "frequency": "M",
                  "missings": True,
                  "missing_value": None,
                  "time_multicolumn": False,
                  "time_composed": True}

        CleanSingleColumnTi._clean_time_index(ws, params)

        wb_exp = load_workbook(
            os.path.join(abs_path("expected"), "test_case5.xlsx"))

        # wb.save("test_case5_after_cleaning_index.xlsx")
        self.assertTrue(compare_cells(wb, wb_exp))

    # @unittest.skip("skip")
    def test_forth_time_value_typo(self):

        exp_time = arrow.get(2015, 5, 2)
        max_forth_time = arrow.get(2015, 5, 22)
        curr_time = arrow.get(2015, 7, 2)
        fixed_time = BaseCleanTiStrategy._forth_time_value_typo(curr_time,
                                                                max_forth_time)
        self.assertEqual(exp_time, fixed_time)


# @unittest.skip("skip")
class CleanMultipleColumnsTiConcatTestCase(unittest.TestCase):

    # @unittest.skip("skip")
    def test_clean_time_index_case5b(self):

        wb = load_workbook(
            os.path.join(abs_path("original"), "test_case5b.xlsx"))
        ws = wb.active

        params = {"time_alignment": 0,
                  "time_format": str,
                  "time_header_coord": ["A18", "B18"],
                  "data_starts": 28,
                  "data_ends": 993,
                  "frequency": "M",
                  "missings": True,
                  "missing_value": None,
                  "time_multicolumn": True,
                  "continuity": False,
                  "time_composed": True}

        CleanMultipleColumnsTiConcat._clean_time_index(ws, params)

        wb_exp = load_workbook(
            os.path.join(abs_path("expected"), "test_case5.xlsx"))

        # wb.save("test_case5b_after_cleaning_index.xlsx")
        self.assertTrue(compare_cells(wb, wb_exp))


if __name__ == '__main__':
    nose.run(defaultTest=__name__)
