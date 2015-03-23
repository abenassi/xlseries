#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
test_strategies
----------------------------------

Tests for `strategies` module.
"""

import unittest
import nose
import os
from openpyxl import load_workbook
import pandas as pd

from xlseries.utils import get_data_frames
from xlseries.strategies import ParameterDiscovery
from xlseries.parameters import Parameters
from utils import compare_data_frames, compare_period_ranges


def load_wb_and_data_frame(path):
    """Call test with a test workbook and expected data frame.

    Args:
        path: Relative path where the df xl file is located."""

    def test_decorator(fn):

        base_path = os.path.join(os.path.dirname(__file__), path)

        # load expected data frame
        exp_file_name = parse_t_name(fn.__name__) + "_exp.xlsx"
        exp_file_path = os.path.join(base_path, exp_file_name)
        exp_dfs = get_data_frames(exp_file_path)

        # get result data frame from xlseries for the test xl file
        test_file_name = parse_t_name(fn.__name__) + ".xlsx"
        test_file_path = os.path.join(base_path, test_file_name)
        test_wb = load_workbook(test_file_path)

        def test_decorated(self):
            fn(self, test_wb=test_wb, exp_dfs=exp_dfs)

        test_decorated.__name__ = fn.__name__
        return test_decorated
    return test_decorator


def load_params(path):
    """Load parameters in a test case.

    Args:
        path: Relative path where the params json file is located."""

    def test_decorator(fn):

        base_path = os.path.join(os.path.dirname(__file__), path)

        # parse parameters from json file
        params_file_name = parse_t_name(fn.__name__) + "_params.json"
        params_file_path = os.path.join(base_path, params_file_name)
        params = Parameters(params_file_path)

        def test_decorated(self, test_wb, exp_dfs):
            fn(self, test_wb=test_wb, exp_dfs=exp_dfs, params=params)

        test_decorated.__name__ = fn.__name__
        return test_decorated
    return test_decorator


def parse_t_name(fn_name):
    """Parse the test name from a function name."""
    return "_".join(fn_name.split("_")[:2])


# @unittest.skip("skip")
class ParameterDiscoveryTestCase(unittest.TestCase):

    @load_wb_and_data_frame("cases")
    @load_params("cases")
    # @unittest.skip("skip")
    def test_case1_with_params(self, test_wb, exp_dfs, params):
        """Test the strategy with case1 and providing parameters."""

        # get dfs from the strategy
        strategy_obj = ParameterDiscovery(test_wb, params)
        test_dfs = strategy_obj.get_data_frames()

        for test_df, exp_df in zip(test_dfs, exp_dfs):
            self.assertTrue(compare_data_frames(test_df, exp_df))

    @load_wb_and_data_frame("cases")
    @load_params("cases")
    # @unittest.skip("skip")
    def test_case2_with_params(self, test_wb, exp_dfs, params):
        """Test the strategy with case2 and providing parameters."""

        # get dfs from the strategy
        strategy_obj = ParameterDiscovery(test_wb, params)
        test_dfs = strategy_obj.get_data_frames()

        for test_df, exp_df in zip(test_dfs, exp_dfs):

            msg = "Different index size: " + str(test_df.index.size) + \
                "  " + str(exp_df.index.size)
            assert test_df.index.size == exp_df.index.size, msg

            self.assertTrue(compare_data_frames(test_df, exp_df))

    # @load_wb_and_data_frame("cases")
    # @load_params("cases")
    @unittest.skip("skip")
    def test_case3_with_params(self, test_wb, exp_dfs, params):
        """Test the strategy with case3 and providing parameters."""

        # get dfs from the strategy
        strategy_obj = ParameterDiscovery(test_wb, params)
        test_dfs = strategy_obj.get_data_frames()

        for test_df, exp_df in zip(test_dfs, exp_dfs):
            self.assertTrue(compare_data_frames(test_df, exp_df))


# @unittest.skip("skip")
class PeriodRangeTestCase(unittest.TestCase):

    def test_get_period_ranges(self):

        test_wb = load_workbook("cases/test_case2.xlsx")
        params = Parameters("cases/test_case2_params.json")
        strategy_obj = ParameterDiscovery(test_wb, params)
        ws = strategy_obj.wb.active

        pr_d = pd.period_range("20020304", "20140410", freq="D")
        pr_m = pd.period_range("20020301", "20130301", freq="M")

        period_ranges = strategy_obj._get_period_ranges(ws)

        self.assertTrue(compare_period_ranges(pr_d, period_ranges[0]))
        self.assertTrue(compare_period_ranges(pr_m, period_ranges[1]))


if __name__ == '__main__':
    nose.run(defaultTest=__name__)
